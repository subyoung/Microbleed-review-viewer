from __future__ import annotations

import hashlib
import json
import os
import sqlite3
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

import dataset_config


# Filled from dataset_config at import, and again by ``configure()`` when a
# study's own config.json is loaded.  Mutated in place rather than rebound:
# other modules hold a reference to these objects.
MAIN_SHEET = dataset_config.DEFAULTS["workbook"]["sheet"]
SOURCE_COLUMNS: dict[str, str] = dict(dataset_config.DEFAULTS["workbook"]["columns"])

def configure(config: dict[str, Any] | None = None) -> dict[str, Any]:
    """Point the store at a dataset's shape.

    Called once at start-up.  Everything it sets is read at call time, so a
    module that imported these names before still sees the new values.
    """

    global MAIN_SHEET
    checked = dataset_config.validate(config or {})
    MAIN_SHEET = checked["workbook"]["sheet"]
    SOURCE_COLUMNS.clear()
    SOURCE_COLUMNS.update(checked["workbook"]["columns"])
    for key, entry in checked["sequences"].items():
        spec = MODALITY_SPECS[key]
        spec["label"] = entry["label"]
        spec["short"] = entry["short"]
        spec["expected"] = f"*{entry['suffix']}"
        spec["suffixes"] = (entry["suffix"],)
        spec["required"] = entry["required"]
        spec["segmentable"] = entry["segmentable"]
    return checked


def required_modalities() -> tuple[str, ...]:
    """The sequences a case must have before it counts as readable."""

    return tuple(key for key, spec in MODALITY_SPECS.items() if spec.get("required", True))


MODALITY_SPECS: dict[str, dict[str, Any]] = {
    "qsm": {
        "label": "QSM",
        "short": "QSM",
        "expected": "*_chi_nSFCR+0_Avg_AffineRestored.nii.gz",
        "suffixes": ("_chi_nSFCR+0_Avg_AffineRestored.nii.gz",),
        "required": True,
        "segmentable": True,
    },
    "swi": {
        "label": "SWI",
        "short": "SWI",
        "expected": "*_GRE4_SWI_AffineRestored.nii.gz",
        "suffixes": ("_GRE4_SWI_AffineRestored.nii.gz",),
        "required": True,
        "segmentable": True,
    },
    "mip": {
        "label": "SWI MIP",
        "short": "MIP",
        "expected": "*_GRE4_SWI_axiMIP2_AffineRestored.nii.gz",
        "suffixes": ("_GRE4_SWI_axiMIP2_AffineRestored.nii.gz",),
        # Optional: a projection helps you spot a lesion and does nothing
        # after that, so a study without one is not an incomplete study.
        "required": False,
        # And nothing may be segmented on it -- see the note in the viewer.
        "segmentable": False,
    },
}


# The structured half of a verdict.  Free-form strings in the database, so a
# study can change its vocabulary without a migration; these are what the
# viewer offers.  The mimics are the ones the microbleed rating scales (MARS,
# BOMBS) tell a reader to exclude, which is what makes a "no" analysable.
CERTAINTY_CHOICES = ("definite", "probable", "possible")
MIMIC_CHOICES = (
    "vessel",
    "calcification",
    "cavernous malformation",
    "partial volume",
    "artefact",
    "iron deposition",
    "haemorrhagic lesion",
    "other",
)


class SourceReadError(RuntimeError):
    pass


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str, sort_keys=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


# How long a connection waits for a lock.  A synced folder or a virus scanner
# can hold the file for seconds, and a reader saving a verdict would rather
# wait than lose it.
DEFAULT_BUSY_TIMEOUT_MS = 30_000

_thread_state = threading.local()


def set_busy_timeout_ms(milliseconds: int | None) -> None:
    """Set how long *this thread's* connections wait for a lock.

    Background writes -- the operation log, the session state -- are the ones
    nobody is waiting for, so they should give up quickly rather than hold a
    worker thread for half a minute and turn shutting down into a choice
    between hanging and killing a thread mid-commit.  ``None`` restores the
    default.
    """

    _thread_state.busy_timeout_ms = milliseconds


def busy_timeout_ms() -> int:
    return int(getattr(_thread_state, "busy_timeout_ms", None) or DEFAULT_BUSY_TIMEOUT_MS)


def connect(db_path: Path) -> sqlite3.Connection:
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    timeout_ms = busy_timeout_ms()
    connection = sqlite3.connect(str(db_path), timeout=timeout_ms / 1000.0)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    connection.execute(f"PRAGMA busy_timeout = {timeout_ms}")
    try:
        connection.execute("PRAGMA journal_mode = WAL")
    except sqlite3.DatabaseError:
        # A shared/synced folder may not support WAL. The app remains usable
        # with the default rollback journal in that environment.
        pass
    return connection


def _schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS source_microbleeds (
            source_row INTEGER PRIMARY KEY,
            target_id TEXT NOT NULL UNIQUE,
            case_id TEXT NOT NULL,
            slicecount INTEGER,
            atlasregion TEXT,
            ras_l REAL NOT NULL,
            ras_p REAL NOT NULL,
            ras_s REAL NOT NULL,
            ras_verified INTEGER,
            source_2d_qsm INTEGER,
            source_swi INTEGER,
            source_3d_qsm INTEGER,
            source_readers TEXT,
            source_verify INTEGER,
            source_need_adjudicate TEXT,
            source_comments TEXT,
            imported_at TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_source_case
            ON source_microbleeds(case_id);

        CREATE TABLE IF NOT EXISTS case_inventory (
            case_id TEXT PRIMARY KEY,
            folder_path TEXT,
            folder_exists INTEGER NOT NULL DEFAULT 0,
            qsm_path TEXT,
            swi_path TEXT,
            mip_path TEXT,
            qsm_status TEXT NOT NULL,
            swi_status TEXT NOT NULL,
            mip_status TEXT NOT NULL,
            scanned_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS reader_profiles (
            reader_id TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            last_seen_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS reader_sessions (
            session_id TEXT PRIMARY KEY,
            reader_id TEXT NOT NULL,
            review_round INTEGER NOT NULL,
            started_at TEXT NOT NULL,
            last_seen_at TEXT NOT NULL,
            closed_at TEXT,
            status TEXT NOT NULL,
            resumed_from_session_id TEXT,
            last_case_id TEXT,
            last_target_id TEXT,
            last_modality TEXT,
            last_axial INTEGER,
            last_coronal INTEGER,
            last_sagittal INTEGER,
            filters_json TEXT,
            UNIQUE(reader_id, review_round)
        );

        CREATE INDEX IF NOT EXISTS idx_sessions_reader
            ON reader_sessions(reader_id, started_at DESC);

        -- ``ras_*`` is this reader's correction of where the finding sits.
        -- NULL means the reader accepts the source coordinate; the source row
        -- itself is never modified, so every version stays inspectable.
        CREATE TABLE IF NOT EXISTS review_annotations (
            review_id INTEGER PRIMARY KEY AUTOINCREMENT,
            target_id TEXT NOT NULL,
            case_id TEXT NOT NULL,
            reader_id TEXT NOT NULL,
            review_round INTEGER NOT NULL,
            verify INTEGER,
            comment TEXT,
            ras_l REAL,
            ras_p REAL,
            ras_s REAL,
            -- How far the coordinate this reader stands behind sits from the
            -- nearest local intensity extremum, measured when it was saved.
            -- A coordinate several millimetres from anything focal is one to
            -- look at again: either the point is off, or there is no lesion
            -- there.  NULL when it could not be measured.
            snap_mm REAL,
            -- Yes/no alone cannot be tabulated into anything: a study needs to
            -- know how sure the reader was, and when they said no, what they
            -- thought it was instead.  Free text in ``comment`` carries that
            -- today and no analysis can read it.
            certainty TEXT,
            mimic TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(target_id, reader_id, review_round)
        );

        CREATE INDEX IF NOT EXISTS idx_reviews_target
            ON review_annotations(target_id, reader_id, review_round);

        CREATE TABLE IF NOT EXISTS manual_annotations (
            manual_id TEXT PRIMARY KEY,
            target_id TEXT NOT NULL UNIQUE,
            case_id TEXT NOT NULL,
            ras_l REAL NOT NULL,
            ras_p REAL NOT NULL,
            ras_s REAL NOT NULL,
            atlasregion TEXT,
            initial_note TEXT,
            created_by TEXT NOT NULL,
            review_round INTEGER NOT NULL,
            created_at TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_manual_case
            ON manual_annotations(case_id, created_at);

        -- One segmentation per finding per reader.  The mask itself lives in a
        -- NIfTI beside the review database, one file per case and reader, with
        -- ``label_value`` picking this finding out of it.  ``generated_from``
        -- records which sequence the assisted growing used, because SWI and
        -- QSM can disagree and the reader needs to know which one they saw.
        CREATE TABLE IF NOT EXISTS roi_labels (
            roi_id INTEGER PRIMARY KEY AUTOINCREMENT,
            target_id TEXT NOT NULL,
            case_id TEXT NOT NULL,
            reader_id TEXT NOT NULL,
            review_round INTEGER NOT NULL,
            label_value INTEGER NOT NULL,
            path TEXT NOT NULL,
            voxel_count INTEGER NOT NULL,
            volume_mm3 REAL NOT NULL,
            centroid_l REAL,
            centroid_p REAL,
            centroid_s REAL,
            generated_from TEXT,
            -- How the mask was made, and under what settings.  A volume in
            -- mm3 is not reproducible without them: the same number can come
            -- from a hand-drawn mask or from region growing at a threshold
            -- nobody recorded.  NULL where it does not apply -- a mask drawn
            -- entirely by hand has no threshold.
            method TEXT,
            sensitivity REAL,
            radius_mm REAL,
            updated_at TEXT NOT NULL,
            UNIQUE(target_id, reader_id, review_round)
        );

        CREATE INDEX IF NOT EXISTS idx_roi_case
            ON roi_labels(case_id, reader_id, review_round);

        -- ``origin_*`` name the row this one was copied from during a merge:
        -- the store it came from and its log_id there.  A store's own entries
        -- leave both NULL, and SQLite treats NULLs in a unique index as
        -- distinct, so only merged rows are deduplicated -- which is what makes
        -- re-running a merge a no-op instead of doubling the log.
        CREATE TABLE IF NOT EXISTS operation_log (
            log_id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT,
            reader_id TEXT,
            review_round INTEGER,
            event_type TEXT NOT NULL,
            case_id TEXT,
            target_id TEXT,
            details_json TEXT,
            created_at TEXT NOT NULL,
            origin_store_id TEXT,
            origin_log_id INTEGER
        );

        CREATE INDEX IF NOT EXISTS idx_log_case
            ON operation_log(case_id, created_at DESC);
        """
    )
    # ``idx_log_origin`` is created in ``_migrate``, not here: on a database
    # that predates the origin columns the CREATE TABLE above is a no-op, so
    # indexing those columns in the same script fails before the migration has
    # had a chance to add them -- which would make every existing store
    # unopenable.
    _migrate(connection)


def _migrate(connection: sqlite3.Connection) -> None:
    """Additive migrations, so an older store keeps working after an upgrade."""

    columns = {
        str(row["name"])
        for row in connection.execute("PRAGMA table_info(review_annotations)")
    }
    for column in ("ras_l", "ras_p", "ras_s", "snap_mm"):
        if column not in columns:
            connection.execute(f"ALTER TABLE review_annotations ADD COLUMN {column} REAL")
    for column in ("certainty", "mimic"):
        if column not in columns:
            connection.execute(f"ALTER TABLE review_annotations ADD COLUMN {column} TEXT")
    roi_columns = {str(row["name"]) for row in connection.execute("PRAGMA table_info(roi_labels)")}
    if roi_columns:
        for column in ("centroid_l", "centroid_p", "centroid_s", "sensitivity", "radius_mm"):
            if column not in roi_columns:
                connection.execute(f"ALTER TABLE roi_labels ADD COLUMN {column} REAL")
        if "method" not in roi_columns:
            # Left NULL on existing rows on purpose: we do not know how they
            # were made, and guessing would be worse than admitting it.
            connection.execute("ALTER TABLE roi_labels ADD COLUMN method TEXT")
    log_columns = {str(row["name"]) for row in connection.execute("PRAGMA table_info(operation_log)")}
    if log_columns:
        for column, kind in (("origin_store_id", "TEXT"), ("origin_log_id", "INTEGER")):
            if column not in log_columns:
                connection.execute(f"ALTER TABLE operation_log ADD COLUMN {column} {kind}")
        # Unconditional, and only once the columns are certain to exist.
        connection.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_log_origin "
            "ON operation_log(origin_store_id, origin_log_id)"
        )
    # Every store carries an identity of its own, so a merge can tell which of
    # its rows already came from where.  Generated once and never changed.
    if get_meta(connection, "store_id") is None:
        set_meta(connection, "store_id", uuid.uuid4().hex)
    connection.commit()


def get_meta(connection: sqlite3.Connection, key: str) -> str | None:
    row = connection.execute("SELECT value FROM meta WHERE key = ?", (key,)).fetchone()
    return str(row["value"]) if row else None


def set_meta(connection: sqlite3.Connection, key: str, value: Any) -> None:
    connection.execute(
        "INSERT INTO meta(key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, str(value)),
    )


def _safe_value(row: dict[str, Any], key: str) -> Any:
    if not key:
        # The configuration says this study does not have that column.
        return None
    value = row.get(key)
    if value == "":
        return None
    return value


def _import_source_rows(connection: sqlite3.Connection, source_xlsx: Path) -> int:
    parsed = _read_source_rows(Path(source_xlsx))
    imported_at = utc_now()
    connection.execute("BEGIN")
    try:
        connection.executemany(
            """
            INSERT INTO source_microbleeds(
                source_row, target_id, case_id, slicecount, atlasregion,
                ras_l, ras_p, ras_s, ras_verified, source_2d_qsm,
                source_swi, source_3d_qsm, source_readers, source_verify,
                source_need_adjudicate, source_comments, imported_at
            ) VALUES (
                :source_row, :target_id, :case_id, :slicecount, :atlasregion,
                :ras_l, :ras_p, :ras_s, :ras_verified, :source_2d_qsm,
                :source_swi, :source_3d_qsm, :source_readers, :source_verify,
                :source_need_adjudicate, :source_comments, :imported_at
            )
            """,
            [{**row, "imported_at": imported_at} for row in parsed],
        )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    return len(parsed)


SOURCE_METADATA_COLUMNS = (
    "case_id",
    "slicecount",
    "atlasregion",
    "ras_l",
    "ras_p",
    "ras_s",
    "ras_verified",
    "source_2d_qsm",
    "source_swi",
    "source_3d_qsm",
    "source_readers",
    "source_verify",
    "source_need_adjudicate",
    "source_comments",
)


def _read_source_rows(source_xlsx: Path) -> list[dict[str, Any]]:
    """Parse the workbook into one dict per finding, keyed by Excel row."""

    try:
        workbook = load_workbook(source_xlsx, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - platform-specific file errors
        raise SourceReadError(
            f"Could not read source workbook: {source_xlsx}. "
            "Make sure the file is readable by the Python process."
        ) from exc

    # Read everything, then let go.  openpyxl's read-only mode keeps the file
    # open until it is closed, which on Windows means the viewer holds a lock
    # on the reader's own workbook for as long as it runs -- Excel then refuses
    # to save it.  Nothing below needs the handle.
    try:
        if MAIN_SHEET not in workbook.sheetnames:
            raise SourceReadError(
                f"Required sheet '{MAIN_SHEET}' was not found in {source_xlsx}. "
                f"The file has: {', '.join(workbook.sheetnames)}. "
                "Set workbook.sheet in config.json if yours is named differently."
            )
        rows = list(workbook[MAIN_SHEET].iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise SourceReadError(f"Source sheet '{MAIN_SHEET}' is empty.")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    column = SOURCE_COLUMNS
    required = {column[key] for key in ("case_id", "ras_l", "ras_p", "ras_s", "verify")}
    missing = sorted(required - set(headers))
    if missing:
        # Name what was there as well: nine times out of ten the column exists
        # under a slightly different heading, and the fix is one line of
        # config.json rather than an edit to the workbook.
        raise SourceReadError(
            f"Missing required source columns: {', '.join(missing)}. "
            f"The sheet has: {', '.join(name for name in headers if name)}. "
            "Set workbook.columns in config.json if yours are named differently."
        )

    parsed: list[dict[str, Any]] = []
    for excel_row, values in enumerate(rows[1:], start=2):
        if not any(value is not None for value in values):
            continue
        row = dict(zip(headers, values))
        case_id = _safe_value(row, column["case_id"])
        if not case_id:
            continue
        try:
            ras_l = float(row[column["ras_l"]])
            ras_p = float(row[column["ras_p"]])
            ras_s = float(row[column["ras_s"]])
        except (TypeError, ValueError) as exc:
            raise SourceReadError(f"Invalid RAS coordinate on Excel row {excel_row}.") from exc
        parsed.append(
            {
                "source_row": excel_row,
                "target_id": f"source:{excel_row}",
                "case_id": str(case_id),
                "slicecount": _safe_value(row, column["slice"]),
                "atlasregion": _safe_value(row, column["region"]),
                "ras_l": ras_l,
                "ras_p": ras_p,
                "ras_s": ras_s,
                "ras_verified": _safe_value(row, column["ras_verified"]),
                "source_2d_qsm": _safe_value(row, column["flag_2d_qsm"]),
                "source_swi": _safe_value(row, column["flag_swi"]),
                "source_3d_qsm": _safe_value(row, column["flag_3d_qsm"]),
                "source_readers": _safe_value(row, column["readers"]),
                "source_verify": _safe_value(row, column["verify"]),
                "source_need_adjudicate": _safe_value(row, column["adjudicate"]),
                "source_comments": _safe_value(row, column["comments"]),
            }
        )
    return parsed


def reimport_source(db_path: Path, source_xlsx: Path) -> dict[str, Any]:
    """Bring workbook changes into an existing store without losing reviews.

    Rows are matched on their Excel row number.  New rows are added, changed
    source metadata is refreshed, and reader annotations are never touched --
    a finding that has been reviewed keeps its review even if the workbook row
    was edited.  Rows that disappeared from the workbook are reported but not
    deleted, because a reader may already have written about them.
    """

    source_xlsx = Path(source_xlsx)
    parsed = _read_source_rows(source_xlsx)
    connection = connect(Path(db_path))
    try:
        _schema(connection)
        existing = {
            int(row["source_row"]): dict(row)
            for row in connection.execute("SELECT * FROM source_microbleeds")
        }
        now = utc_now()
        added: list[int] = []
        updated: list[int] = []
        unchanged = 0
        connection.execute("BEGIN")
        try:
            for row in parsed:
                source_row = int(row["source_row"])
                current = existing.get(source_row)
                if current is None:
                    connection.execute(
                        """
                        INSERT INTO source_microbleeds(
                            source_row, target_id, case_id, slicecount, atlasregion,
                            ras_l, ras_p, ras_s, ras_verified, source_2d_qsm,
                            source_swi, source_3d_qsm, source_readers, source_verify,
                            source_need_adjudicate, source_comments, imported_at
                        ) VALUES (
                            :source_row, :target_id, :case_id, :slicecount, :atlasregion,
                            :ras_l, :ras_p, :ras_s, :ras_verified, :source_2d_qsm,
                            :source_swi, :source_3d_qsm, :source_readers, :source_verify,
                            :source_need_adjudicate, :source_comments, :imported_at
                        )
                        """,
                        {**row, "imported_at": now},
                    )
                    added.append(source_row)
                    continue
                differences = {
                    column
                    for column in SOURCE_METADATA_COLUMNS
                    if _normalise_source_value(current.get(column)) != _normalise_source_value(row.get(column))
                }
                if not differences:
                    unchanged += 1
                    continue
                assignments = ", ".join(f"{column} = :{column}" for column in SOURCE_METADATA_COLUMNS)
                connection.execute(
                    f"UPDATE source_microbleeds SET {assignments}, imported_at = :imported_at "
                    "WHERE source_row = :source_row",
                    {**row, "imported_at": now},
                )
                updated.append(source_row)
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        seen = {int(row["source_row"]) for row in parsed}
        removed = sorted(set(existing) - seen)
        try:
            set_meta(connection, "source_snapshot_hash", sha256_file(source_xlsx))
            set_meta(connection, "source_reimported_at", now)
            connection.commit()
        except OSError:
            pass
        return {
            "added": len(added),
            "updated": len(updated),
            "unchanged": unchanged,
            "removed_from_workbook": len(removed),
            "removed_rows": removed,
            "total_in_workbook": len(parsed),
        }
    finally:
        connection.close()


def _normalise_source_value(value: Any) -> Any:
    """Compare workbook values the way a human would read them."""

    if value is None or value == "":
        return None
    if isinstance(value, float):
        return round(value, 9)
    if isinstance(value, int):
        return float(value)
    text = str(value).strip()
    try:
        return round(float(text), 9)
    except ValueError:
        return text


def _find_modality_file(folder: Path, modality: str) -> Path | None:
    # The review coordinates are only trusted with the AffineRestored
    # products.  Do not fall back to similarly named non-restored NIfTIs.
    spec = MODALITY_SPECS[modality]
    names = sorted(path.name for path in folder.iterdir() if path.is_file())
    for suffix in spec["suffixes"]:
        matches = [name for name in names if name.endswith(suffix)]
        if matches:
            return folder / matches[0]
    return None


def _file_status(folder_exists: bool, path: Path | None) -> str:
    if not folder_exists:
        return "missing_folder"
    return "available" if path else "missing"


def refresh_inventory(connection: sqlite3.Connection, data_root: Path) -> dict[str, int]:
    data_root = Path(data_root)
    case_ids = [row["case_id"] for row in connection.execute("SELECT DISTINCT case_id FROM source_microbleeds")]
    counts = {"complete": 0, "partial": 0, "all_missing": 0, "missing_folder": 0}
    scanned_at = utc_now()
    for case_id in case_ids:
        folder = data_root / case_id
        folder_exists = folder.is_dir()
        paths = {
            modality: _find_modality_file(folder, modality) if folder_exists else None
            for modality in MODALITY_SPECS
        }
        available_count = sum(path is not None for path in paths.values())
        # "Complete" means every *required* sequence is there.  A study whose
        # MIP is optional should not have every case marked partial for the
        # want of a projection nobody reads.
        required = required_modalities()
        if not folder_exists:
            inventory_status = "missing_folder"
        elif available_count == 0:
            inventory_status = "all_missing"
        elif all(paths.get(key) is not None for key in required):
            inventory_status = "complete"
        else:
            inventory_status = "partial"
        counts[inventory_status] += 1
        connection.execute(
            """
            INSERT INTO case_inventory(
                case_id, folder_path, folder_exists, qsm_path, swi_path, mip_path,
                qsm_status, swi_status, mip_status, scanned_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(case_id) DO UPDATE SET
                folder_path = excluded.folder_path,
                folder_exists = excluded.folder_exists,
                qsm_path = excluded.qsm_path,
                swi_path = excluded.swi_path,
                mip_path = excluded.mip_path,
                qsm_status = excluded.qsm_status,
                swi_status = excluded.swi_status,
                mip_status = excluded.mip_status,
                scanned_at = excluded.scanned_at
            """,
            (
                case_id,
                str(folder),
                int(folder_exists),
                str(paths["qsm"]) if paths["qsm"] else None,
                str(paths["swi"]) if paths["swi"] else None,
                str(paths["mip"]) if paths["mip"] else None,
                _file_status(folder_exists, paths["qsm"]),
                _file_status(folder_exists, paths["swi"]),
                _file_status(folder_exists, paths["mip"]),
                scanned_at,
            ),
        )
    connection.commit()
    return counts


def initialize_store(source_xlsx: Path, data_root: Path, db_path: Path) -> dict[str, Any]:
    source_xlsx = Path(source_xlsx)
    data_root = Path(data_root)
    db_path = Path(db_path)
    connection = connect(db_path)
    try:
        _schema(connection)
        stored_hash = get_meta(connection, "source_snapshot_hash")
        source_count = connection.execute("SELECT COUNT(*) AS n FROM source_microbleeds").fetchone()["n"]
        imported = False
        source_unavailable = False
        if source_count == 0:
            if not source_xlsx.exists():
                raise SourceReadError(f"Source workbook was not found: {source_xlsx}")
            try:
                current_hash = sha256_file(source_xlsx)
            except OSError as exc:
                raise SourceReadError(
                    f"Could not access the source workbook for first-time import: {source_xlsx}"
                ) from exc
            source_count = _import_source_rows(connection, source_xlsx)
            set_meta(connection, "source_snapshot_hash", current_hash)
            set_meta(connection, "source_snapshot_path", source_xlsx)
            set_meta(connection, "source_initialized_at", utc_now())
            set_meta(connection, "source_sheet", MAIN_SHEET)
            connection.commit()
            imported = True
        else:
            current_hash: str | None = None
            if source_xlsx.exists():
                try:
                    current_hash = sha256_file(source_xlsx)
                except OSError:
                    source_unavailable = True
            else:
                source_unavailable = True
            if current_hash is not None and stored_hash is None:
                set_meta(connection, "source_snapshot_hash", current_hash)
                connection.commit()
            elif current_hash is not None and stored_hash != current_hash:
                # Preserve the imported source snapshot. The UI can report that
                # the original workbook changed and an explicit re-import is required.
                set_meta(connection, "source_latest_hash", current_hash)
                set_meta(connection, "source_latest_seen_at", utc_now())
                connection.commit()
        inventory = refresh_inventory(connection, data_root)
        case_count = connection.execute(
            "SELECT COUNT(DISTINCT case_id) AS n FROM source_microbleeds"
        ).fetchone()["n"]
        return {
            "source_count": int(source_count),
            "case_count": int(case_count),
            "imported": imported,
            "source_changed": (
                not source_unavailable
                and stored_hash is not None
                and current_hash is not None
                and stored_hash != current_hash
            ),
            "source_unavailable": source_unavailable,
            "inventory": inventory,
            "db_path": str(db_path),
        }
    finally:
        connection.close()


def refresh_inventory_store(db_path: Path, data_root: Path) -> dict[str, int]:
    connection = connect(Path(db_path))
    try:
        _schema(connection)
        return refresh_inventory(connection, Path(data_root))
    finally:
        connection.close()


def distance_mm(first: Iterable[float] | None, second: Iterable[float] | None) -> float | None:
    """Straight-line distance between two RAS points, in millimetres."""

    if first is None or second is None:
        return None
    a = [float(value) for value in first]
    b = [float(value) for value in second]
    if len(a) != 3 or len(b) != 3:
        return None
    return float(sum((x - y) ** 2 for x, y in zip(a, b)) ** 0.5)


def _case_file_status(row: sqlite3.Row | dict[str, Any]) -> str:
    if not row["folder_exists"]:
        return "missing_folder"
    paths = [row["qsm_path"], row["swi_path"], row["mip_path"]]
    if not any(paths):
        return "all_missing"
    # Optional sequences do not hold a case back; see refresh_inventory.
    if all(row[f"{key}_path"] for key in required_modalities()):
        return "complete"
    return "partial"


def list_cases(db_path: Path, reader_id: str, review_round: int) -> list[dict[str, Any]]:
    connection = connect(db_path)
    try:
        rows = connection.execute(
            """
            -- Progress is about every finding a reader sees in the panel, and
            -- that includes the ones they added themselves.  The source-only
            -- columns stay source-only, so the workbook filters keep meaning
            -- what they say.
            WITH findings AS (
                SELECT case_id, target_id, source_verify, source_need_adjudicate, 1 AS is_source
                FROM source_microbleeds
                UNION ALL
                SELECT case_id, target_id, NULL, NULL, 0 FROM manual_annotations
            ),
            -- One row per finding that at least two readers decided
            -- differently, whichever rounds they used.
            disagreements AS (
                SELECT target_id
                FROM review_annotations
                WHERE verify IS NOT NULL
                GROUP BY target_id
                HAVING COUNT(DISTINCT reader_id) > 1 AND COUNT(DISTINCT verify) > 1
            )
            SELECT
                c.*,
                COUNT(f.target_id) AS finding_count,
                COALESCE(SUM(f.is_source), 0) AS source_count,
                COALESCE(SUM(CASE WHEN r.review_id IS NOT NULL
                    AND (r.verify IS NOT NULL OR NULLIF(TRIM(r.comment), '') IS NOT NULL)
                    THEN 1 ELSE 0 END), 0) AS reviewed_count,
                COALESCE(SUM(CASE WHEN r.verify = 1 THEN 1 ELSE 0 END), 0) AS reader_verified_count,
                COALESCE(SUM(CASE WHEN r.verify = 0 THEN 1 ELSE 0 END), 0) AS reader_not_verified_count,
                COALESCE(SUM(CASE WHEN f.is_source = 1
                    AND (f.source_verify IS NULL OR f.source_verify != 1)
                    THEN 1 ELSE 0 END), 0) AS source_unverified_count,
                COALESCE(SUM(CASE WHEN NULLIF(TRIM(r.comment), '') IS NOT NULL
                    THEN 1 ELSE 0 END), 0) AS reader_comment_count,
                COALESCE(SUM(CASE WHEN NULLIF(TRIM(f.source_need_adjudicate), '') IS NOT NULL
                    THEN 1 ELSE 0 END), 0) AS adjudication_count,
                -- Findings where the readers in this database have actually
                -- disagreed.  The source sheet's own adjudication notes answer
                -- a different question: this one is about the work in progress.
                COALESCE(SUM(CASE WHEN d.target_id IS NOT NULL THEN 1 ELSE 0 END), 0)
                    AS disagreement_count
            FROM case_inventory c
            LEFT JOIN findings f ON f.case_id = c.case_id
            LEFT JOIN review_annotations r
                ON r.target_id = f.target_id
                AND r.reader_id = ?
                AND r.review_round = ?
            LEFT JOIN disagreements d ON d.target_id = f.target_id
            GROUP BY c.case_id
            ORDER BY c.case_id
            """,
            (reader_id, review_round),
        ).fetchall()
        cases: list[dict[str, Any]] = []
        for row in rows:
            item = dict(row)
            item["source_count"] = int(item["source_count"] or 0)
            item["finding_count"] = int(item["finding_count"] or 0)
            item["reviewed_count"] = int(item["reviewed_count"] or 0)
            item["reader_verified_count"] = int(item["reader_verified_count"] or 0)
            item["reader_not_verified_count"] = int(item["reader_not_verified_count"] or 0)
            item["reader_unverified_count"] = item["finding_count"] - item["reader_verified_count"]
            item["source_unverified_count"] = int(item["source_unverified_count"] or 0)
            item["reader_comment_count"] = int(item["reader_comment_count"] or 0)
            item["adjudication_count"] = int(item["adjudication_count"] or 0)
            item["disagreement_count"] = int(item["disagreement_count"] or 0)
            item["file_status"] = _case_file_status(item)
            # Progress is measured by findings the reader has actually decided
            # on.  Counting only ``verify = 1`` left a case whose findings were
            # all deliberately rejected looking permanently unreviewed.
            item["reader_review_status"] = (
                "Unreviewed"
                if item["reviewed_count"] == 0
                else "Reviewed"
                if item["reviewed_count"] >= item["finding_count"]
                else "In progress"
            )
            cases.append(item)
        return cases
    finally:
        connection.close()


def get_case(db_path: Path, case_id: str) -> dict[str, Any] | None:
    connection = connect(db_path)
    try:
        row = connection.execute(
            "SELECT * FROM case_inventory WHERE case_id = ?", (case_id,)
        ).fetchone()
        return dict(row) if row else None
    finally:
        connection.close()


def list_targets(
    db_path: Path,
    case_id: str,
    reader_id: str,
    review_round: int,
) -> list[dict[str, Any]]:
    connection = connect(db_path)
    try:
        def reports_for(target_id: str) -> list[dict[str, Any]]:
            report_rows = connection.execute(
                """
                SELECT reader_id, review_round, verify, comment,
                       ras_l, ras_p, ras_s, snap_mm, certainty, mimic,
                       created_at, updated_at
                FROM review_annotations
                WHERE target_id = ?
                ORDER BY updated_at DESC, reader_id, review_round DESC
                """,
                (target_id,),
            ).fetchall()
            return [dict(report) for report in report_rows]

        def variants_for(reports: list[dict[str, Any]], source_ras: tuple[float, float, float]) -> list[dict[str, Any]]:
            """Every recorded opinion of where this finding sits."""

            variants = [
                {
                    "key": "source",
                    "reader_id": None,
                    "review_round": None,
                    "ras": source_ras,
                    "moved_mm": 0.0,
                    "updated_at": None,
                }
            ]
            seen: set[str] = set()
            for report in reports:
                if report.get("ras_l") is None:
                    continue
                reader = str(report["reader_id"])
                if reader in seen:
                    continue
                seen.add(reader)
                ras = (float(report["ras_l"]), float(report["ras_p"]), float(report["ras_s"]))
                variants.append(
                    {
                        "key": reader,
                        "reader_id": reader,
                        "review_round": report.get("review_round"),
                        "ras": ras,
                        "moved_mm": distance_mm(source_ras, ras),
                        "updated_at": report.get("updated_at"),
                    }
                )
            return variants

        source_rows = connection.execute(
            """
            SELECT s.*, r.verify AS reader_verify, r.comment AS reader_comment,
                   r.certainty AS reader_certainty, r.mimic AS reader_mimic,
                   r.ras_l AS reader_ras_l, r.ras_p AS reader_ras_p, r.ras_s AS reader_ras_s,
                   r.updated_at AS reader_updated_at
            FROM source_microbleeds s
            LEFT JOIN review_annotations r
                ON r.target_id = s.target_id
                AND r.reader_id = ?
                AND r.review_round = ?
            WHERE s.case_id = ?
            ORDER BY s.source_row
            """,
            (reader_id, review_round, case_id),
        ).fetchall()
        manual_rows = connection.execute(
            """
            SELECT m.*, r.verify AS reader_verify, r.comment AS reader_comment,
                   r.certainty AS reader_certainty, r.mimic AS reader_mimic,
                   r.ras_l AS reader_ras_l, r.ras_p AS reader_ras_p, r.ras_s AS reader_ras_s,
                   r.updated_at AS reader_updated_at
            FROM manual_annotations m
            LEFT JOIN review_annotations r
                ON r.target_id = m.target_id
                AND r.reader_id = ?
                AND r.review_round = ?
            WHERE m.case_id = ?
            ORDER BY m.created_at
            """,
            (reader_id, review_round, case_id),
        ).fetchall()

        try:
            roi_rows = {
                str(row["target_id"]): dict(row)
                for row in connection.execute(
                    """
                    SELECT * FROM roi_labels
                    WHERE case_id = ? AND reader_id = ? AND review_round = ?
                    """,
                    (case_id, reader_id, review_round),
                )
            }
        except sqlite3.OperationalError:
            # A store written before segmentation existed; it gains the table
            # the next time it is opened through ``initialize_store``.
            roi_rows = {}

        def finalise(item: dict[str, Any]) -> dict[str, Any]:
            source_ras = (
                float(item["ras_l"]),
                float(item["ras_p"]),
                float(item["ras_s"]),
            )
            reports = reports_for(str(item["target_id"]))
            reader_ras = None
            if item.get("reader_ras_l") is not None:
                reader_ras = (
                    float(item["reader_ras_l"]),
                    float(item["reader_ras_p"]),
                    float(item["reader_ras_s"]),
                )
            item.update(
                {
                    "ras": source_ras,
                    "source_ras": source_ras,
                    "reader_ras": reader_ras,
                    # Where this finding is for this reader: their own
                    # correction when they made one, the source otherwise.
                    "effective_ras": reader_ras or source_ras,
                    "reader_moved_mm": distance_mm(source_ras, reader_ras) if reader_ras else None,
                    "reader_reports": reports,
                    "position_variants": variants_for(reports, source_ras),
                    "roi": roi_rows.get(str(item["target_id"])),
                }
            )
            return item

        targets: list[dict[str, Any]] = []
        for row in source_rows:
            item = dict(row)
            item.update(
                {
                    "origin": "Source",
                    "label": f"Source #{item['source_row']}",
                    "manual_id": None,
                    "manual_note": None,
                }
            )
            targets.append(finalise(item))
        for row in manual_rows:
            item = dict(row)
            item.update(
                {
                    "origin": "Manual",
                    "label": f"Manual #{item['manual_id'][:8]}",
                    "target_id": item["target_id"],
                    "source_row": None,
                    "atlasregion": item.get("atlasregion"),
                    "source_verify": None,
                    "source_readers": None,
                    "source_need_adjudicate": None,
                    "source_comments": None,
                    "manual_id": item["manual_id"],
                    "manual_note": item.get("initial_note"),
                }
            )
            targets.append(finalise(item))
        return targets
    finally:
        connection.close()


def register_reader(db_path: Path, reader_id: str) -> None:
    reader_id = reader_id.strip()
    now = utc_now()
    connection = connect(db_path)
    try:
        connection.execute(
            """
            INSERT INTO reader_profiles(reader_id, display_name, created_at, last_seen_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(reader_id) DO UPDATE SET
                display_name = excluded.display_name,
                last_seen_at = excluded.last_seen_at
            """,
            (reader_id, reader_id, now, now),
        )
        connection.commit()
    finally:
        connection.close()


def get_resume_candidate(db_path: Path, reader_id: str) -> dict[str, Any] | None:
    connection = connect(db_path)
    try:
        row = connection.execute(
            """
            SELECT * FROM reader_sessions
            WHERE reader_id = ? AND status IN ('open', 'closed')
            ORDER BY started_at DESC
            LIMIT 1
            """,
            (reader_id,),
        ).fetchone()
        return dict(row) if row else None
    finally:
        connection.close()


def list_reader_rounds(db_path: Path, reader_id: str) -> list[dict[str, Any]]:
    """Every review round this reader has, newest first, with its progress.

    A reader who started a new round by accident needs to be able to go back to
    the earlier one, so the choice is made from the rounds themselves rather
    than from "the latest session".
    """

    connection = connect(db_path)
    try:
        rounds: list[dict[str, Any]] = []
        for row in connection.execute(
            """
            SELECT review_round,
                   MIN(started_at) AS started_at,
                   MAX(last_seen_at) AS last_seen_at,
                   COUNT(*) AS session_count
            FROM reader_sessions
            WHERE reader_id = ?
            GROUP BY review_round
            ORDER BY review_round DESC
            """,
            (reader_id,),
        ).fetchall():
            entry = dict(row)
            review_round = int(entry["review_round"])
            latest = connection.execute(
                """
                SELECT session_id, status, last_case_id, last_target_id, last_modality
                FROM reader_sessions
                WHERE reader_id = ? AND review_round = ?
                ORDER BY last_seen_at DESC, started_at DESC
                LIMIT 1
                """,
                (reader_id, review_round),
            ).fetchone()
            entry.update(dict(latest) if latest else {})
            entry["reviewed_count"] = int(
                connection.execute(
                    """
                    SELECT COUNT(*) AS n FROM review_annotations
                    WHERE reader_id = ? AND review_round = ?
                      AND (verify IS NOT NULL OR NULLIF(TRIM(comment), '') IS NOT NULL)
                    """,
                    (reader_id, review_round),
                ).fetchone()["n"]
                or 0
            )
            try:
                entry["roi_count"] = int(
                    connection.execute(
                        "SELECT COUNT(*) AS n FROM roi_labels WHERE reader_id = ? AND review_round = ?",
                        (reader_id, review_round),
                    ).fetchone()["n"]
                    or 0
                )
            except sqlite3.OperationalError:
                entry["roi_count"] = 0
            rounds.append(entry)
        return rounds
    finally:
        connection.close()


def _new_session_id() -> str:
    return uuid.uuid4().hex


def start_new_session(
    db_path: Path,
    reader_id: str,
    resumed_from_session_id: str | None = None,
) -> dict[str, Any]:
    register_reader(db_path, reader_id)
    now = utc_now()
    session_id = _new_session_id()
    connection = connect(db_path)
    try:
        previous = connection.execute(
            "SELECT MAX(review_round) AS n FROM reader_sessions WHERE reader_id = ?",
            (reader_id,),
        ).fetchone()
        review_round = int(previous["n"] or 0) + 1
        connection.execute(
            """
            UPDATE reader_sessions SET status = 'superseded', closed_at = ?
            WHERE reader_id = ? AND status = 'open'
            """,
            (now, reader_id),
        )
        connection.execute(
            """
            INSERT INTO reader_sessions(
                session_id, reader_id, review_round, started_at, last_seen_at,
                closed_at, status, resumed_from_session_id
            ) VALUES (?, ?, ?, ?, ?, NULL, 'open', ?)
            """,
            (session_id, reader_id, review_round, now, now, resumed_from_session_id),
        )
        connection.commit()
        return {
            "session_id": session_id,
            "reader_id": reader_id,
            "review_round": review_round,
            "started_at": now,
            "status": "open",
        }
    finally:
        connection.close()


def resume_session(db_path: Path, session_id: str) -> dict[str, Any]:
    now = utc_now()
    connection = connect(db_path)
    try:
        connection.execute(
            "UPDATE reader_sessions SET status = 'open', closed_at = NULL, last_seen_at = ? WHERE session_id = ?",
            (now, session_id),
        )
        connection.commit()
        row = connection.execute(
            "SELECT * FROM reader_sessions WHERE session_id = ?", (session_id,)
        ).fetchone()
        if not row:
            raise ValueError(f"Unknown session: {session_id}")
        return dict(row)
    finally:
        connection.close()


def close_session(db_path: Path, session_id: str) -> None:
    now = utc_now()
    connection = connect(db_path)
    try:
        connection.execute(
            "UPDATE reader_sessions SET status = 'closed', closed_at = ?, last_seen_at = ? WHERE session_id = ?",
            (now, now, session_id),
        )
        connection.commit()
    finally:
        connection.close()


def save_session_state(db_path: Path, session_id: str, state: dict[str, Any]) -> None:
    connection = connect(db_path)
    try:
        connection.execute(
            """
            UPDATE reader_sessions SET
                last_seen_at = ?, last_case_id = ?, last_target_id = ?,
                last_modality = ?, last_axial = ?, last_coronal = ?,
                last_sagittal = ?, filters_json = ?
            WHERE session_id = ?
            """,
            (
                utc_now(),
                state.get("case_id"),
                state.get("target_id"),
                state.get("modality"),
                state.get("axial"),
                state.get("coronal"),
                state.get("sagittal"),
                json_dumps(state.get("filters", {})),
                session_id,
            ),
        )
        connection.commit()
    finally:
        connection.close()


def log_event(
    db_path: Path,
    event_type: str,
    *,
    session_id: str | None = None,
    reader_id: str | None = None,
    review_round: int | None = None,
    case_id: str | None = None,
    target_id: str | None = None,
    details: dict[str, Any] | None = None,
) -> None:
    connection = connect(db_path)
    try:
        connection.execute(
            """
            INSERT INTO operation_log(
                session_id, reader_id, review_round, event_type, case_id,
                target_id, details_json, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session_id,
                reader_id,
                review_round,
                event_type,
                case_id,
                target_id,
                json_dumps(details or {}),
                utc_now(),
            ),
        )
        connection.commit()
    finally:
        connection.close()


def save_review(
    db_path: Path,
    *,
    target_id: str,
    case_id: str,
    reader_id: str,
    review_round: int,
    verify: int | None,
    comment: str | None,
    corrected_ras: Iterable[float] | None = None,
    session_id: str | None = None,
    snap_mm: float | None = None,
    certainty: str | None = None,
    mimic: str | None = None,
) -> None:
    """Record this reader's opinion of a finding.

    ``corrected_ras`` is where this reader thinks the finding actually is.
    ``None`` means they accept the source coordinate.  The source row is never
    edited, so a correction adds a version rather than replacing one.

    ``snap_mm`` is how far that coordinate sits from the nearest local
    intensity extremum, measured while the image was still in memory.  It is
    the cheapest available answer to "is this point actually on anything?".

    ``certainty`` and ``mimic`` are the structured half of the verdict: how
    sure the reader was, and what they thought it was instead when they said
    no.  Both are free-form strings so the study can change its vocabulary
    without a migration; the viewer offers ``CERTAINTY_CHOICES`` and
    ``MIMIC_CHOICES``.
    """

    if verify not in (None, 0, 1):
        raise ValueError("verify must be None, 0, or 1")
    comment = comment.strip() if comment and comment.strip() else None
    ras: tuple[float, float, float] | tuple[None, None, None]
    if corrected_ras is None:
        ras = (None, None, None)
    else:
        values = tuple(float(value) for value in corrected_ras)
        if len(values) != 3:
            raise ValueError("A corrected coordinate requires three RAS values.")
        ras = values
    now = utc_now()
    connection = connect(db_path)
    try:
        existing = connection.execute(
            """
            SELECT verify, comment, ras_l, ras_p, ras_s, updated_at FROM review_annotations
            WHERE target_id = ? AND reader_id = ? AND review_round = ?
            """,
            (target_id, reader_id, review_round),
        ).fetchone()
        connection.execute(
            """
            INSERT INTO review_annotations(
                target_id, case_id, reader_id, review_round, verify, comment,
                ras_l, ras_p, ras_s, snap_mm, certainty, mimic, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(target_id, reader_id, review_round) DO UPDATE SET
                verify = excluded.verify,
                comment = excluded.comment,
                ras_l = excluded.ras_l,
                ras_p = excluded.ras_p,
                ras_s = excluded.ras_s,
                snap_mm = excluded.snap_mm,
                certainty = excluded.certainty,
                mimic = excluded.mimic,
                updated_at = excluded.updated_at
            """,
            (
                target_id, case_id, reader_id, review_round, verify, comment, *ras,
                float(snap_mm) if snap_mm is not None else None,
                (certainty or None), (mimic or None), now, now,
            ),
        )
        connection.commit()
    finally:
        connection.close()
    log_event(
        db_path,
        "review_saved",
        session_id=session_id,
        reader_id=reader_id,
        review_round=review_round,
        case_id=case_id,
        target_id=target_id,
        details={
            "before": dict(existing) if existing else None,
            "after": {"verify": verify, "comment": comment, "corrected_ras": ras if corrected_ras is not None else None},
        },
    )


def add_manual_annotation(
    db_path: Path,
    *,
    case_id: str,
    ras: Iterable[float],
    reader_id: str,
    review_round: int,
    atlasregion: str | None = None,
    initial_note: str | None = None,
    session_id: str | None = None,
) -> str:
    values = tuple(float(value) for value in ras)
    if len(values) != 3:
        raise ValueError("A manual annotation requires three RAS values.")
    manual_id = uuid.uuid4().hex
    target_id = f"manual:{manual_id}"
    now = utc_now()
    connection = connect(db_path)
    try:
        connection.execute(
            """
            INSERT INTO manual_annotations(
                manual_id, target_id, case_id, ras_l, ras_p, ras_s,
                atlasregion, initial_note, created_by, review_round, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                manual_id,
                target_id,
                case_id,
                values[0],
                values[1],
                values[2],
                atlasregion.strip() if atlasregion and atlasregion.strip() else None,
                initial_note.strip() if initial_note and initial_note.strip() else None,
                reader_id,
                review_round,
                now,
            ),
        )
        connection.commit()
    finally:
        connection.close()
    log_event(
        db_path,
        "manual_microbleed_added",
        session_id=session_id,
        reader_id=reader_id,
        review_round=review_round,
        case_id=case_id,
        target_id=target_id,
        details={"ras": values, "atlasregion": atlasregion, "initial_note": initial_note},
    )
    return target_id


# Two distinct findings of one case never come within 2.03 mm in this dataset:
# over all 1016 within-case pairs, none is closer than 2 mm, one sits at 2.03
# and three under 4.  So a point within a millimetre of an existing finding is
# that finding, and one within three is close enough to be worth a question --
# a reader adding a genuine second lesion will see that question about once in
# a thousand pairs, and one adding a duplicate will see it every time.
SAME_FINDING_MM = 1.0
NEARBY_FINDING_MM = 3.0


def findings_near(
    db_path: Path,
    case_id: str,
    ras: Iterable[float],
    within_mm: float = NEARBY_FINDING_MM,
) -> list[dict[str, Any]]:
    """Findings of this case already close to a coordinate, nearest first."""

    point = tuple(float(value) for value in ras)
    if len(point) != 3:
        raise ValueError("A coordinate needs three RAS values.")
    connection = connect(Path(db_path))
    try:
        rows: list[dict[str, Any]] = []
        for table, label_column in (
            ("source_microbleeds", "source_row"),
            ("manual_annotations", "manual_id"),
        ):
            for row in connection.execute(
                f"SELECT * FROM {table} WHERE case_id = ?",  # noqa: S608 - fixed names
                (case_id,),
            ):
                item = dict(row)
                distance = distance_mm(point, (item["ras_l"], item["ras_p"], item["ras_s"]))
                if distance is None or distance > float(within_mm):
                    continue
                manual = table == "manual_annotations"
                rows.append(
                    {
                        "target_id": str(item["target_id"]),
                        "origin": "Manual" if manual else "Source",
                        "label": (
                            f"Manual #{str(item['manual_id'])[:8]}"
                            if manual
                            else f"Source #{item['source_row']}"
                        ),
                        "distance_mm": distance,
                        "ras": (item["ras_l"], item["ras_p"], item["ras_s"]),
                        "created_by": item.get("created_by") if manual else None,
                        "_sort": item[label_column],
                    }
                )
        rows.sort(key=lambda item: item["distance_mm"])
        for item in rows:
            item.pop("_sort", None)
        return rows
    finally:
        connection.close()


def manual_deletion_blockers(db_path: Path, target_id: str, reader_id: str) -> list[str]:
    """Reasons this finding cannot be removed, in words a reader can act on.

    Separated from the deletion itself so the window can grey a button out and
    explain why without opening anything, and so the rules are testable
    without a dialog.
    """

    connection = connect(Path(db_path))
    try:
        row = connection.execute(
            "SELECT created_by FROM manual_annotations WHERE target_id = ?", (str(target_id),)
        ).fetchone()
        if row is None:
            return [
                "Only a finding somebody added by hand can be removed; this one "
                "comes from the findings workbook."
            ]
        reasons: list[str] = []
        owner = str(row["created_by"])
        if owner != str(reader_id):
            reasons.append(f"{owner} added this finding, so only they can remove it.")
        others = sorted(
            {
                str(item["reader_id"])
                for item in connection.execute(
                    "SELECT reader_id FROM review_annotations WHERE target_id = ?",
                    (str(target_id),),
                )
                if str(item["reader_id"]) != str(reader_id)
            }
        )
        if others:
            reasons.append(
                f"{', '.join(others)} already reviewed it; removing it would delete their work."
            )
        try:
            segmenters = sorted(
                {
                    str(item["reader_id"])
                    for item in connection.execute(
                        "SELECT reader_id FROM roi_labels WHERE target_id = ?", (str(target_id),)
                    )
                    if str(item["reader_id"]) != str(reader_id)
                }
            )
        except sqlite3.OperationalError:
            segmenters = []
        if segmenters:
            reasons.append(f"{', '.join(segmenters)} already segmented it.")
        return reasons
    finally:
        connection.close()


def delete_manual_annotation(
    db_path: Path,
    *,
    target_id: str,
    reader_id: str,
    session_id: str | None = None,
) -> dict[str, Any]:
    """Remove a finding this reader added, and their own work on it.

    The reader's verdict and mask go with it: they describe a finding that no
    longer exists, and leaving them would put rows in the results table with
    nothing to join to.  Anything belonging to another reader stops the
    deletion instead -- see :func:`manual_deletion_blockers`.
    """

    blockers = manual_deletion_blockers(db_path, target_id, reader_id)
    if blockers:
        raise ValueError(" ".join(blockers))
    connection = connect(Path(db_path))
    try:
        row = connection.execute(
            "SELECT case_id, ras_l, ras_p, ras_s FROM manual_annotations WHERE target_id = ?",
            (str(target_id),),
        ).fetchone()
        case_id = str(row["case_id"]) if row else None
        ras = (row["ras_l"], row["ras_p"], row["ras_s"]) if row else None
        reviews = connection.execute(
            "DELETE FROM review_annotations WHERE target_id = ?", (str(target_id),)
        ).rowcount
        try:
            segmentations = connection.execute(
                "DELETE FROM roi_labels WHERE target_id = ?", (str(target_id),)
            ).rowcount
        except sqlite3.OperationalError:
            segmentations = 0
        connection.execute(
            "DELETE FROM manual_annotations WHERE target_id = ?", (str(target_id),)
        )
        connection.commit()
    finally:
        connection.close()
    log_event(
        db_path,
        "manual_microbleed_removed",
        session_id=session_id,
        reader_id=reader_id,
        case_id=case_id,
        target_id=str(target_id),
        details={"ras": ras, "reviews": int(reviews or 0), "segmentations": int(segmentations or 0)},
    )
    return {
        "target_id": str(target_id),
        "case_id": case_id,
        "reviews": int(reviews or 0),
        "segmentations": int(segmentations or 0),
    }


def _agreement(verdicts: list[int]) -> str:
    """How the readers who recorded a verdict compare."""

    decided = [value for value in verdicts if value in (0, 1)]
    if not decided:
        return "no verdict"
    if len(decided) == 1:
        return "single reader"
    if all(value == 1 for value in decided):
        return "agree yes"
    if all(value == 0 for value in decided):
        return "agree no"
    return "disagreement"


def collect_export_rows(db_path: Path) -> tuple[list[str], list[dict[str, Any]], list[dict[str, Any]]]:
    """Build the analysis table: one row per finding, one column set per reader."""

    connection = connect(Path(db_path))
    try:
        readers = [
            str(row["reader_id"])
            for row in connection.execute(
                "SELECT reader_id FROM review_annotations "
                "UNION SELECT reader_id FROM roi_labels ORDER BY reader_id"
            )
        ]
        rois: dict[str, dict[str, dict[str, Any]]] = {}
        for row in connection.execute("SELECT * FROM roi_labels"):
            item = dict(row)
            rois.setdefault(str(item["target_id"]), {})[str(item["reader_id"])] = item
        reviews: dict[str, dict[str, dict[str, Any]]] = {}
        long_rows: list[dict[str, Any]] = []
        for row in connection.execute(
            """
            SELECT target_id, case_id, reader_id, review_round, verify, comment,
                   ras_l, ras_p, ras_s, snap_mm, certainty, mimic, created_at, updated_at
            FROM review_annotations
            ORDER BY target_id, reader_id, review_round
            """
        ):
            item = dict(row)
            long_rows.append(item)
            # The latest round is what the analysis should use.
            reviews.setdefault(str(item["target_id"]), {})[str(item["reader_id"])] = item

        findings: list[dict[str, Any]] = []
        for row in connection.execute("SELECT * FROM source_microbleeds ORDER BY source_row"):
            item = dict(row)
            findings.append(
                {
                    "case_id": item["case_id"],
                    "target_id": item["target_id"],
                    "origin": "Source",
                    "source_row": item["source_row"],
                    "atlasregion": item["atlasregion"],
                    "ras_l": item["ras_l"],
                    "ras_p": item["ras_p"],
                    "ras_s": item["ras_s"],
                    "source_verify": item["source_verify"],
                    "source_readers": item["source_readers"],
                    "source_need_adjudicate": item["source_need_adjudicate"],
                    "source_comments": item["source_comments"],
                    "created_by": None,
                    "initial_note": None,
                }
            )
        for row in connection.execute("SELECT * FROM manual_annotations ORDER BY created_at"):
            item = dict(row)
            findings.append(
                {
                    "case_id": item["case_id"],
                    "target_id": item["target_id"],
                    "origin": "Manual",
                    "source_row": None,
                    "atlasregion": item["atlasregion"],
                    "ras_l": item["ras_l"],
                    "ras_p": item["ras_p"],
                    "ras_s": item["ras_s"],
                    "source_verify": None,
                    "source_readers": None,
                    "source_need_adjudicate": None,
                    "source_comments": None,
                    "created_by": item["created_by"],
                    "initial_note": item["initial_note"],
                }
            )

        for finding in findings:
            per_target = reviews.get(str(finding["target_id"]), {})
            source_ras = (finding["ras_l"], finding["ras_p"], finding["ras_s"])
            verdicts: list[int] = []
            moved = 0
            for reader in readers:
                review = per_target.get(reader)
                verify = review.get("verify") if review else None
                finding[f"{reader} · verify"] = verify
                finding[f"{reader} · comment"] = review.get("comment") if review else None
                finding[f"{reader} · certainty"] = review.get("certainty") if review else None
                finding[f"{reader} · mimic"] = review.get("mimic") if review else None
                finding[f"{reader} · round"] = review.get("review_round") if review else None
                # Full corrected coordinates live in the reader_reports sheet;
                # the wide sheet carries how far the reader moved the finding.
                corrected = None
                if review and review.get("ras_l") is not None:
                    corrected = (review["ras_l"], review["ras_p"], review["ras_s"])
                    moved += 1
                finding[f"{reader} · moved_mm"] = (
                    round(distance_mm(source_ras, corrected), 3) if corrected else None
                )
                # How far the coordinate they stand behind is from anything
                # focal: a quality flag on the point itself, not on the verdict.
                finding[f"{reader} · snap_mm"] = (
                    round(float(review["snap_mm"]), 3)
                    if review and review.get("snap_mm") is not None
                    else None
                )
                # The coordinate this reader stands behind: their correction if
                # they made one, otherwise the source coordinate they accepted.
                # Blank when the reader never reviewed the finding at all.
                final = corrected if corrected else (source_ras if review else None)
                finding[f"{reader} · final_l"] = final[0] if final else None
                finding[f"{reader} · final_p"] = final[1] if final else None
                finding[f"{reader} · final_s"] = final[2] if final else None
                roi = rois.get(str(finding["target_id"]), {}).get(reader)
                # Without the label value the NIfTI cannot be read back: one file
                # holds every finding of the case, one integer each.
                finding[f"{reader} · roi_label"] = int(roi["label_value"]) if roi else None
                finding[f"{reader} · roi_mm3"] = round(float(roi["volume_mm3"]), 3) if roi else None
                finding[f"{reader} · roi_from"] = roi.get("generated_from") if roi else None
                finding[f"{reader} · roi_file"] = Path(str(roi["path"])).name if roi else None
                finding[f"{reader} · updated"] = review.get("updated_at") if review else None
                if verify in (0, 1):
                    verdicts.append(int(verify))
            finding["readers_with_verdict"] = len(verdicts)
            finding["verdict_yes"] = sum(1 for value in verdicts if value == 1)
            finding["verdict_no"] = sum(1 for value in verdicts if value == 0)
            finding["readers_who_moved_it"] = moved
            finding["readers_who_segmented_it"] = sum(
                1 for reader in readers if rois.get(str(finding["target_id"]), {}).get(reader)
            )
            finding["agreement"] = _agreement(verdicts)
        return readers, findings, long_rows
    finally:
        connection.close()


SEGMENTATION_COLUMNS = [
    "case_id",
    "target_id",
    "reader_id",
    "review_round",
    "label_value",
    "label_file",
    "verify",
    "comment",
    "voxel_count",
    "volume_mm3",
    "centroid_l",
    "centroid_p",
    "centroid_s",
    "generated_from",
    "method",
    "sensitivity",
    "radius_mm",
    "source_l",
    "source_p",
    "source_s",
    "final_l",
    "final_p",
    "final_s",
    "moved_mm",
    "updated_at",
]


def collect_segmentation_rows(db_path: Path) -> list[dict[str, Any]]:
    """One row per segmented finding, keyed by the integer inside the NIfTI.

    A label file holds every finding of one case for one reader and round, each
    under its own ``label_value``. Those findings do not share a verdict -- a
    reader can segment two blobs and call one of them a microbleed and the other
    not -- so the verdict, the comment and the coordinate travel with the label
    value rather than with the file.
    """

    connection = connect(Path(db_path))
    try:
        sources: dict[str, dict[str, Any]] = {}
        for table in ("source_microbleeds", "manual_annotations"):
            for row in connection.execute(f"SELECT * FROM {table}"):  # noqa: S608 - fixed names
                item = dict(row)
                sources[str(item["target_id"])] = item
        reviews: dict[tuple[str, str, int], dict[str, Any]] = {}
        for row in connection.execute("SELECT * FROM review_annotations"):
            item = dict(row)
            reviews[
                (str(item["target_id"]), str(item["reader_id"]), int(item["review_round"]))
            ] = item

        rows: list[dict[str, Any]] = []
        try:
            roi_rows = connection.execute(
                "SELECT * FROM roi_labels ORDER BY case_id, reader_id, review_round, label_value"
            ).fetchall()
        except sqlite3.OperationalError:
            roi_rows = []
        for row in roi_rows:
            roi = dict(row)
            target_id = str(roi["target_id"])
            source = sources.get(target_id, {})
            source_ras = (source.get("ras_l"), source.get("ras_p"), source.get("ras_s"))
            review = reviews.get((target_id, str(roi["reader_id"]), int(roi["review_round"])))
            corrected = None
            if review and review.get("ras_l") is not None:
                corrected = (review["ras_l"], review["ras_p"], review["ras_s"])
            final = corrected or (source_ras if None not in source_ras else (None, None, None))
            rows.append(
                {
                    "case_id": roi["case_id"],
                    "target_id": target_id,
                    "reader_id": roi["reader_id"],
                    "review_round": roi["review_round"],
                    "label_value": roi["label_value"],
                    "label_file": Path(str(roi["path"])).name,
                    "verify": review.get("verify") if review else None,
                    "comment": review.get("comment") if review else None,
                    "voxel_count": roi["voxel_count"],
                    "volume_mm3": round(float(roi["volume_mm3"]), 3),
                    # Where the voxels are, which is not the same as where the
                    # finding is marked: a reader can move the point after
                    # drawing, or draw a lesion that is not centred on it.
                    "centroid_l": roi.get("centroid_l"),
                    "centroid_p": roi.get("centroid_p"),
                    "centroid_s": roi.get("centroid_s"),
                    "generated_from": roi.get("generated_from"),
                    # How it was made, so two volumes in this column can be
                    # compared -- or knowingly not compared.
                    "method": roi.get("method"),
                    "sensitivity": roi.get("sensitivity"),
                    "radius_mm": roi.get("radius_mm"),
                    "source_l": source_ras[0],
                    "source_p": source_ras[1],
                    "source_s": source_ras[2],
                    "final_l": final[0],
                    "final_p": final[1],
                    "final_s": final[2],
                    "moved_mm": (
                        round(distance_mm(source_ras, corrected), 3)
                        if corrected and None not in source_ras
                        else None
                    ),
                    "updated_at": roi["updated_at"],
                }
            )
        return rows
    finally:
        connection.close()


AGREEMENT_COLUMNS = [
    "case_id",
    "target_id",
    "reader_a",
    "reader_b",
    "dice",
    "centroid_mm",
    "volume_a_mm3",
    "volume_b_mm3",
    "volume_ratio",
]


def collect_agreement_rows(db_path: Path) -> list[dict[str, Any]]:
    """How closely two readers' segmentations of one finding agree.

    Dice and the distance between the two centroids are the numbers an
    inter-rater reliability section reports, and neither can be worked out
    from the results table as it stands: the volumes are there, but two masks
    of the same size can sit in different places.

    Only findings that at least two readers segmented are compared, and each
    label file is read once, so the cost is proportional to the double-read
    work actually done rather than to the size of the study.
    """

    connection = connect(Path(db_path))
    try:
        try:
            rows = [dict(row) for row in connection.execute("SELECT * FROM roi_labels")]
        except sqlite3.OperationalError:
            return []
    finally:
        connection.close()

    by_target: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        if int(row["voxel_count"] or 0) <= 0:
            continue
        by_target.setdefault(str(row["target_id"]), []).append(row)

    import numpy as np

    cache: dict[str, Any] = {}

    def mask_of(entry: dict[str, Any]):
        path = str(entry["path"])
        if path not in cache:
            try:
                import nibabel as nib

                cache[path] = np.asanyarray(nib.load(path).dataobj)
            except Exception:
                # A label file that has moved or cannot be read is reported as
                # a missing comparison, not as a crash in the middle of an
                # export the reader asked for.
                cache[path] = None
        volume = cache[path]
        if volume is None:
            return None
        return volume == int(entry["label_value"])

    results: list[dict[str, Any]] = []
    for target_id, entries in sorted(by_target.items()):
        # One entry per reader: their latest round.
        latest: dict[str, dict[str, Any]] = {}
        for entry in sorted(entries, key=lambda item: str(item["updated_at"])):
            latest[str(entry["reader_id"])] = entry
        readers = sorted(latest)
        if len(readers) < 2:
            continue
        for index, reader_a in enumerate(readers):
            for reader_b in readers[index + 1:]:
                first, second = latest[reader_a], latest[reader_b]
                mask_a, mask_b = mask_of(first), mask_of(second)
                dice = None
                if mask_a is not None and mask_b is not None and mask_a.shape == mask_b.shape:
                    total = int(mask_a.sum()) + int(mask_b.sum())
                    dice = (
                        round(2.0 * float(np.logical_and(mask_a, mask_b).sum()) / total, 4)
                        if total
                        else None
                    )
                centroid = distance_mm(
                    (first.get("centroid_l"), first.get("centroid_p"), first.get("centroid_s")),
                    (second.get("centroid_l"), second.get("centroid_p"), second.get("centroid_s")),
                ) if first.get("centroid_l") is not None and second.get("centroid_l") is not None else None
                volume_a = float(first["volume_mm3"])
                volume_b = float(second["volume_mm3"])
                results.append(
                    {
                        "case_id": first["case_id"],
                        "target_id": target_id,
                        "reader_a": reader_a,
                        "reader_b": reader_b,
                        "dice": dice,
                        "centroid_mm": round(centroid, 3) if centroid is not None else None,
                        "volume_a_mm3": round(volume_a, 3),
                        "volume_b_mm3": round(volume_b, 3),
                        "volume_ratio": (
                            round(max(volume_a, volume_b) / min(volume_a, volume_b), 3)
                            if min(volume_a, volume_b) > 0
                            else None
                        ),
                    }
                )
    return results


def export_reviews(db_path: Path, out_path: Path) -> dict[str, Any]:
    """Write the review results to .xlsx (three sheets) or .csv (plus a sidecar)."""

    out_path = Path(out_path)
    readers, findings, long_rows = collect_export_rows(Path(db_path))
    roi_rows = collect_segmentation_rows(Path(db_path))
    agreement_rows = collect_agreement_rows(Path(db_path))
    base_columns = [
        "case_id",
        "target_id",
        "origin",
        "source_row",
        "atlasregion",
        "ras_l",
        "ras_p",
        "ras_s",
        "source_verify",
        "source_readers",
        "source_need_adjudicate",
        "source_comments",
        "created_by",
        "initial_note",
    ]
    reader_columns = [
        f"{reader} · {field}"
        for reader in readers
        for field in (
            "verify",
            "certainty",
            "mimic",
            "comment",
            "final_l",
            "final_p",
            "final_s",
            "moved_mm",
            "snap_mm",
            "roi_label",
            "roi_mm3",
            "roi_from",
            "roi_file",
            "round",
            "updated",
        )
    ]
    summary_columns = [
        "readers_with_verdict",
        "verdict_yes",
        "verdict_no",
        "readers_who_moved_it",
        "readers_who_segmented_it",
        "agreement",
    ]
    columns = base_columns + reader_columns + summary_columns

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.suffix.lower() == ".csv":
        import csv

        with out_path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            for finding in findings:
                writer.writerow(finding)
        # A csv holds one table, and the label-value map must not be the thing
        # that gets dropped, so it goes in a file beside it.
        roi_path = out_path.with_name(f"{out_path.stem}_segmentations.csv")
        with roi_path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=SEGMENTATION_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            for row in roi_rows:
                writer.writerow(row)
    else:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "findings"
        sheet.append(columns)
        for finding in findings:
            sheet.append([finding.get(column) for column in columns])
        sheet.freeze_panes = "A2"

        long_sheet = workbook.create_sheet("reader_reports")
        long_columns = [
            "case_id",
            "target_id",
            "reader_id",
            "review_round",
            "verify",
            "comment",
            "ras_l",
            "ras_p",
            "ras_s",
            # How far this coordinate is from anything focal, so a reviewer can
            # sort the sheet and find the points that sit on nothing.
            "snap_mm",
            "certainty",
            "mimic",
            "created_at",
            "updated_at",
        ]
        long_sheet.append(long_columns)
        for row in long_rows:
            long_sheet.append([row.get(column) for column in long_columns])
        long_sheet.freeze_panes = "A2"

        roi_sheet = workbook.create_sheet("segmentations")
        roi_sheet.append(SEGMENTATION_COLUMNS)
        for row in roi_rows:
            roi_sheet.append([row.get(column) for column in SEGMENTATION_COLUMNS])
        roi_sheet.freeze_panes = "A2"

        if agreement_rows:
            agreement_sheet = workbook.create_sheet("segmentation_agreement")
            agreement_sheet.append(AGREEMENT_COLUMNS)
            for row in agreement_rows:
                agreement_sheet.append([row.get(column) for column in AGREEMENT_COLUMNS])
            agreement_sheet.freeze_panes = "A2"
        workbook.save(out_path)

    return {
        "path": str(out_path),
        "findings": len(findings),
        "readers": len(readers),
        "reader_reports": len(long_rows),
        "segmentations": len(roi_rows),
        "segmentation_pairs": len(agreement_rows),
        "disagreements": sum(1 for item in findings if item["agreement"] == "disagreement"),
    }


MERGE_TABLES = ("reader_profiles", "reader_sessions", "review_annotations", "manual_annotations")


def _merge_rois(
    target: sqlite3.Connection,
    source: sqlite3.Connection,
    target_db: Path,
    source_db: Path,
    round_map: dict[tuple[str, int], int],
) -> int:
    """Carry one store's segmentations, and their mask files, into another.

    A ``roi_labels`` row is only half of a segmentation: the voxels live in a
    NIfTI beside the source database, one file per case, reader and round.
    Copying the row without the file leaves a record pointing at a path that
    may not exist on the machine doing the merge, so the file is copied into
    the target's own ``labels`` tree and the row is rewritten to point there.
    """

    import shutil

    try:
        rows = source.execute(
            "SELECT * FROM roi_labels ORDER BY case_id, reader_id, review_round, label_value"
        ).fetchall()
    except sqlite3.OperationalError:  # a store written before segmentation existed
        return 0

    copied: dict[tuple[str, str, int], Path] = {}
    added = 0
    for row in rows:
        roi = dict(row)
        roi.pop("roi_id", None)
        reader_id = str(roi["reader_id"])
        old_round = int(roi["review_round"])
        new_round = int(round_map.get((reader_id, old_round), old_round))
        roi["review_round"] = new_round
        case_id = str(roi["case_id"])

        destination = copied.get((case_id, reader_id, new_round))
        if destination is None:
            destination = label_path(target_db, case_id, reader_id, new_round)
            origin = Path(str(roi["path"]))
            if not origin.is_file():
                # The row named a file that is not here.  Fall back to where the
                # source store would have kept it, in case the database was moved
                # together with its labels but the recorded path is stale.
                origin = label_path(source_db, case_id, reader_id, old_round)
            if origin.is_file() and origin.resolve() != destination.resolve():
                destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(origin, destination)
            copied[(case_id, reader_id, new_round)] = destination
        roi["path"] = str(destination)

        existing = target.execute(
            """
            SELECT updated_at FROM roi_labels
            WHERE target_id = ? AND reader_id = ? AND review_round = ?
            """,
            (roi["target_id"], reader_id, new_round),
        ).fetchone()
        if existing is not None and str(existing["updated_at"]) >= str(roi["updated_at"]):
            continue
        columns = list(roi)
        target.execute(
            f"INSERT INTO roi_labels({', '.join(columns)}) "
            f"VALUES ({', '.join(':' + name for name in columns)}) "
            "ON CONFLICT(target_id, reader_id, review_round) DO UPDATE SET "
            + ", ".join(
                f"{name} = excluded.{name}"
                for name in columns
                if name not in ("target_id", "reader_id", "review_round")
            ),
            roi,
        )
        added += 1
    return added


def merge_stores(target_db: Path, source_dbs: Iterable[Path]) -> dict[str, Any]:
    """Fold per-reader review databases into one aggregate database.

    Each reader can work in their own SQLite file -- the safe arrangement when
    the shared folder is synchronised rather than a real server -- and their
    work is combined afterwards.  A review round that would collide with one
    already in the target is renumbered rather than overwritten, so no reader's
    work is ever replaced by another's.
    """

    target = connect(Path(target_db))
    summary: dict[str, Any] = {
        "sources": [],
        "reviews_added": 0,
        "reviews_skipped": 0,
        "manual_added": 0,
        "sessions_added": 0,
        "rounds_renumbered": 0,
        "rois_added": 0,
    }
    try:
        _schema(target)
        for source_path in source_dbs:
            source_path = Path(source_path)
            source = connect(source_path)
            per_source = {
                "path": str(source_path),
                "reviews_added": 0,
                "reviews_skipped": 0,
                "manual_added": 0,
                "sessions_added": 0,
                "rounds_renumbered": 0,
                "rois_added": 0,
            }
            try:
                _schema(source)
                known_sessions = {
                    str(row["session_id"])
                    for row in target.execute("SELECT session_id FROM reader_sessions")
                }
                next_round: dict[str, int] = {}
                for row in target.execute(
                    "SELECT reader_id, MAX(review_round) AS n FROM reader_sessions GROUP BY reader_id"
                ):
                    next_round[str(row["reader_id"])] = int(row["n"] or 0)

                round_map: dict[tuple[str, int], int] = {}
                for row in source.execute("SELECT * FROM reader_sessions ORDER BY started_at"):
                    session = dict(row)
                    reader_id = str(session["reader_id"])
                    old_round = int(session["review_round"])
                    if str(session["session_id"]) in known_sessions:
                        # Already merged in an earlier run; keep its numbering.
                        existing = target.execute(
                            "SELECT review_round FROM reader_sessions WHERE session_id = ?",
                            (session["session_id"],),
                        ).fetchone()
                        round_map[(reader_id, old_round)] = int(existing["review_round"])
                        continue
                    taken = target.execute(
                        "SELECT 1 FROM reader_sessions WHERE reader_id = ? AND review_round = ?",
                        (reader_id, old_round),
                    ).fetchone()
                    new_round = old_round
                    if taken:
                        new_round = max(next_round.get(reader_id, 0), old_round) + 1
                        per_source["rounds_renumbered"] += 1
                    next_round[reader_id] = max(next_round.get(reader_id, 0), new_round)
                    round_map[(reader_id, old_round)] = new_round
                    session["review_round"] = new_round
                    columns = list(session)
                    target.execute(
                        f"INSERT INTO reader_sessions({', '.join(columns)}) "
                        f"VALUES ({', '.join(':' + name for name in columns)})",
                        session,
                    )
                    per_source["sessions_added"] += 1

                for row in source.execute("SELECT * FROM reader_profiles"):
                    profile = dict(row)
                    target.execute(
                        """
                        INSERT INTO reader_profiles(reader_id, display_name, created_at, last_seen_at)
                        VALUES (:reader_id, :display_name, :created_at, :last_seen_at)
                        ON CONFLICT(reader_id) DO UPDATE SET
                            last_seen_at = MAX(reader_profiles.last_seen_at, excluded.last_seen_at)
                        """,
                        profile,
                    )

                for row in source.execute("SELECT * FROM manual_annotations"):
                    manual = dict(row)
                    reader_id = str(manual["created_by"])
                    manual["review_round"] = round_map.get(
                        (reader_id, int(manual["review_round"])), int(manual["review_round"])
                    )
                    columns = list(manual)
                    cursor = target.execute(
                        f"INSERT OR IGNORE INTO manual_annotations({', '.join(columns)}) "
                        f"VALUES ({', '.join(':' + name for name in columns)})",
                        manual,
                    )
                    per_source["manual_added"] += int(cursor.rowcount or 0)

                for row in source.execute("SELECT * FROM review_annotations"):
                    review = dict(row)
                    review.pop("review_id", None)
                    reader_id = str(review["reader_id"])
                    review["review_round"] = round_map.get(
                        (reader_id, int(review["review_round"])), int(review["review_round"])
                    )
                    existing = target.execute(
                        """
                        SELECT updated_at FROM review_annotations
                        WHERE target_id = ? AND reader_id = ? AND review_round = ?
                        """,
                        (review["target_id"], reader_id, review["review_round"]),
                    ).fetchone()
                    if existing is not None:
                        if str(existing["updated_at"]) >= str(review["updated_at"]):
                            per_source["reviews_skipped"] += 1
                            continue
                        # The coordinate is part of the opinion, not decoration:
                        # leaving ras_* behind here kept a stale correction next
                        # to a freshly updated verdict.
                        target.execute(
                            """
                            UPDATE review_annotations SET verify = :verify, comment = :comment,
                                ras_l = :ras_l, ras_p = :ras_p, ras_s = :ras_s,
                                snap_mm = :snap_mm, certainty = :certainty,
                                mimic = :mimic, updated_at = :updated_at
                            WHERE target_id = :target_id AND reader_id = :reader_id
                              AND review_round = :review_round
                            """,
                            review,
                        )
                        per_source["reviews_added"] += 1
                        continue
                    columns = list(review)
                    target.execute(
                        f"INSERT INTO review_annotations({', '.join(columns)}) "
                        f"VALUES ({', '.join(':' + name for name in columns)})",
                        review,
                    )
                    per_source["reviews_added"] += 1

                per_source["rois_added"] += _merge_rois(
                    target, source, Path(target_db), source_path, round_map
                )

                source_store_id = get_meta(source, "store_id") or str(source_path)
                for row in source.execute("SELECT * FROM operation_log ORDER BY log_id"):
                    entry = dict(row)
                    log_id = entry.pop("log_id", None)
                    reader_id = str(entry.get("reader_id") or "")
                    if entry.get("review_round") is not None:
                        entry["review_round"] = round_map.get(
                            (reader_id, int(entry["review_round"])), entry["review_round"]
                        )
                    # Merging an already-merged store keeps the original
                    # provenance, so a chain of merges stays idempotent too.
                    entry["origin_store_id"] = entry.get("origin_store_id") or source_store_id
                    entry["origin_log_id"] = entry.get("origin_log_id") or log_id
                    columns = list(entry)
                    target.execute(
                        f"INSERT OR IGNORE INTO operation_log({', '.join(columns)}) "
                        f"VALUES ({', '.join(':' + name for name in columns)})",
                        entry,
                    )
                target.commit()
            finally:
                source.close()
            summary["sources"].append(per_source)
            for key in (
                "reviews_added",
                "reviews_skipped",
                "manual_added",
                "sessions_added",
                "rounds_renumbered",
                "rois_added",
            ):
                summary[key] += per_source[key]
        return summary
    finally:
        target.close()


def label_directory(db_path: Path, reader_id: str) -> Path:
    """Where this reader's segmentations live: beside the review database."""

    safe = "".join(character if character.isalnum() or character in "-_" else "_" for character in reader_id)
    return Path(db_path).parent / "labels" / (safe or "reader")


def label_path(db_path: Path, case_id: str, reader_id: str, review_round: int) -> Path:
    return label_directory(db_path, reader_id) / f"{case_id}_round{int(review_round)}.nii.gz"


def list_rois(db_path: Path, case_id: str, reader_id: str, review_round: int) -> dict[str, dict[str, Any]]:
    connection = connect(db_path)
    try:
        rows = connection.execute(
            """
            SELECT * FROM roi_labels
            WHERE case_id = ? AND reader_id = ? AND review_round = ?
            """,
            (case_id, reader_id, review_round),
        ).fetchall()
        return {str(row["target_id"]): dict(row) for row in rows}
    finally:
        connection.close()


def save_roi(
    db_path: Path,
    *,
    target_id: str,
    case_id: str,
    reader_id: str,
    review_round: int,
    label_value: int,
    path: Path,
    voxel_count: int,
    volume_mm3: float,
    generated_from: str | None,
    centroid_ras: tuple[float, float, float] | None = None,
    session_id: str | None = None,
    method: str | None = None,
    sensitivity: float | None = None,
    radius_mm: float | None = None,
) -> None:
    """Record (or clear) this reader's segmentation of one finding.

    ``method`` is ``"grow"``, ``"brush"`` or ``"grow+brush"``; ``sensitivity``
    and ``radius_mm`` are the settings any growing used.  Without them a
    volume in the results table cannot be reproduced or compared.
    """

    now = utc_now()
    connection = connect(db_path)
    try:
        if int(voxel_count) <= 0:
            connection.execute(
                """
                DELETE FROM roi_labels
                WHERE target_id = ? AND reader_id = ? AND review_round = ?
                """,
                (target_id, reader_id, review_round),
            )
        else:
            connection.execute(
                """
                INSERT INTO roi_labels(
                    target_id, case_id, reader_id, review_round, label_value,
                    path, voxel_count, volume_mm3,
                    centroid_l, centroid_p, centroid_s, generated_from,
                    method, sensitivity, radius_mm, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(target_id, reader_id, review_round) DO UPDATE SET
                    label_value = excluded.label_value,
                    path = excluded.path,
                    voxel_count = excluded.voxel_count,
                    volume_mm3 = excluded.volume_mm3,
                    centroid_l = excluded.centroid_l,
                    centroid_p = excluded.centroid_p,
                    centroid_s = excluded.centroid_s,
                    generated_from = excluded.generated_from,
                    method = excluded.method,
                    sensitivity = excluded.sensitivity,
                    radius_mm = excluded.radius_mm,
                    updated_at = excluded.updated_at
                """,
                (
                    target_id,
                    case_id,
                    reader_id,
                    review_round,
                    int(label_value),
                    str(path),
                    int(voxel_count),
                    float(volume_mm3),
                    float(centroid_ras[0]) if centroid_ras else None,
                    float(centroid_ras[1]) if centroid_ras else None,
                    float(centroid_ras[2]) if centroid_ras else None,
                    generated_from,
                    method,
                    float(sensitivity) if sensitivity is not None else None,
                    float(radius_mm) if radius_mm is not None else None,
                    now,
                ),
            )
        connection.commit()
    finally:
        connection.close()
    log_event(
        db_path,
        "roi_saved",
        session_id=session_id,
        reader_id=reader_id,
        review_round=review_round,
        case_id=case_id,
        target_id=target_id,
        details={
            "voxels": int(voxel_count),
            "volume_mm3": round(float(volume_mm3), 4),
            "generated_from": generated_from,
        },
    )


def recent_case_log(db_path: Path, case_id: str, limit: int = 25) -> list[dict[str, Any]]:
    connection = connect(db_path)
    try:
        rows = connection.execute(
            """
            SELECT * FROM operation_log
            WHERE case_id = ?
            ORDER BY log_id DESC
            LIMIT ?
            """,
            (case_id, int(limit)),
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()
