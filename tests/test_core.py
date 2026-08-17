from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path

import numpy as np

VIEWER_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(VIEWER_DIR))

from imaging import (  # noqa: E402
    ORIENTATION_PRESETS,
    clamp_voxel,
    grow_lesion,
    load_volume,
    save_label_volume,
    opposite_axcode,
    plane_direction_labels,
    preset_axcodes,
    ras_to_voxel,
    robust_window,
    voxel_in_bounds,
    voxel_to_ras,
)
from review_store import (  # noqa: E402
    SEGMENTATION_COLUMNS,
    add_manual_annotation,
    collect_export_rows,
    collect_segmentation_rows,
    connect,
    export_reviews,
    get_case,
    get_resume_candidate,
    initialize_store,
    label_path,
    list_cases,
    list_rois,
    list_targets,
    merge_stores,
    recent_case_log,
    reimport_source,
    save_review,
    save_roi,
    start_new_session,
    resume_session,
)


def _case_with_status(db_path: Path, status: str) -> str:
    """A case whose files are in a particular state, found rather than named."""

    for item in list_cases(db_path, "sample", 1):
        if item["file_status"] == status:
            return str(item["case_id"])
    raise unittest.SkipTest(f"The configured dataset has no {status} case.")


def _sample_case(db_path: Path) -> str:
    """The case the data-backed tests work on.

    Discovered rather than named: the source should not carry a study's
    identifiers, and a test that hard-codes one only runs on one dataset.
    Ordering follows the case queue, so this is the case the viewer opens.
    """

    listing = list_cases(db_path, "sample", 1)
    complete = [item for item in listing if item["file_status"] == "complete"]
    if not complete:
        # These tests read images.  A dataset with none -- the example
        # workbook in examples/, for instance, which ships no MRI -- exercises
        # the store and nothing else, and saying so beats failing.
        raise unittest.SkipTest(
            "The configured dataset has no case with all three sequences."
        )
    return str(complete[0]["case_id"])


class CoreWorkflowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        source = os.environ.get("TEST_SOURCE_XLSX")
        if not source:
            raise unittest.SkipTest("Set TEST_SOURCE_XLSX to a readable copy of the source workbook.")
        cls.source = Path(source)
        cls.data_root = Path(os.environ.get("TEST_DATA_ROOT", VIEWER_DIR.parent / "Data"))
        cls.temp_dir = Path(tempfile.mkdtemp(prefix="microbleed_viewer_test_"))
        cls.db_path = cls.temp_dir / "review.sqlite"
        cls.report = initialize_store(cls.source, cls.data_root, cls.db_path)
        cls.case_id = _sample_case(cls.db_path)

    @classmethod
    def tearDownClass(cls) -> None:
        import shutil

        shutil.rmtree(cls.temp_dir, ignore_errors=True)

    def test_the_import_accounts_for_every_row_and_every_case(self) -> None:
        """The guard is that the parts add up, not that they add up to 434.

        An importer that silently drops rows, or an inventory that loses a
        case between its buckets, is what this is for -- and pinning one
        study's totals only catches it on that study.  On the dataset this
        was written against the figures are 434 findings over 215 cases,
        125 of them complete.
        """

        # Its own store: the class-level one is shared, and a sibling test
        # adds a manual finding to it, which is a finding the workbook never
        # had.  Counting rows against a store somebody else has written to is
        # how a passing test starts depending on the order it runs in.
        folder = Path(tempfile.mkdtemp(prefix="microbleed_counts_test_"))
        try:
            db_path = folder / "review.sqlite"
            report = initialize_store(self.source, self.data_root, db_path)
            inventory = report["inventory"]
            self.assertGreater(report["source_count"], 0, "the workbook imported nothing")
            self.assertGreater(report["case_count"], 0)
            # Every case is in exactly one bucket.
            buckets = ("complete", "partial", "all_missing", "missing_folder")
            self.assertEqual(
                sum(int(inventory[name]) for name in buckets),
                report["case_count"],
                f"the inventory buckets do not add up to the cases: {inventory}",
            )
            # And every workbook row reached the case list.
            cases = list_cases(db_path, "count QA", 1)
            self.assertEqual(len(cases), report["case_count"])
            self.assertEqual(
                sum(int(item["finding_count"]) for item in cases),
                report["source_count"],
                "findings went missing between the workbook and the case list",
            )
        finally:
            import shutil

            shutil.rmtree(folder, ignore_errors=True)

    def test_non_affine_files_are_treated_as_missing(self) -> None:
        complete = get_case(self.db_path, self.case_id)
        self.assertIsNotNone(complete)
        self.assertTrue(complete["qsm_path"])
        self.assertTrue(complete["swi_path"])
        self.assertTrue(complete["mip_path"])

        # A folder that exists but holds no accepted product: other NIfTI
        # variants must not be substituted for the ones the viewer requires.
        present_but_unusable = _case_with_status(self.db_path, "all_missing")
        non_affine_only = get_case(self.db_path, present_but_unusable)
        self.assertIsNotNone(non_affine_only)
        self.assertIsNone(non_affine_only["qsm_path"])
        self.assertIsNone(non_affine_only["swi_path"])
        self.assertIsNone(non_affine_only["mip_path"])
        non_affine_summary = next(
            item for item in list_cases(self.db_path, "File QA", 1)
            if item["case_id"] == present_but_unusable
        )
        self.assertEqual(non_affine_summary["file_status"], "all_missing")

        missing = get_case(self.db_path, _case_with_status(self.db_path, "missing_folder"))
        self.assertIsNotNone(missing)
        self.assertEqual(missing["folder_exists"], 0)
        self.assertIsNone(missing["qsm_path"])

    def test_reader_round_resume_and_restart(self) -> None:
        first = start_new_session(self.db_path, "QA Reader")
        candidates = get_resume_candidate(self.db_path, "QA Reader")
        self.assertEqual(candidates["session_id"], first["session_id"])

        resumed = resume_session(self.db_path, first["session_id"])
        self.assertEqual(resumed["review_round"], 1)
        second = start_new_session(self.db_path, "QA Reader", first["session_id"])
        self.assertEqual(second["review_round"], 2)
        self.assertNotEqual(first["session_id"], second["session_id"])

    def test_reader_review_and_manual_annotation_persist(self) -> None:
        session = start_new_session(self.db_path, "Persistence QA")
        source_targets = list_targets(self.db_path, self.case_id, "Persistence QA", 1)
        self.assertEqual(len(source_targets), 1)
        target = source_targets[0]
        save_review(
            self.db_path,
            target_id=target["target_id"],
            case_id=self.case_id,
            reader_id="Persistence QA",
            review_round=1,
            verify=1,
            comment="QA comment",
            session_id=session["session_id"],
        )
        manual_id = add_manual_annotation(
            self.db_path,
            case_id=self.case_id,
            ras=target["ras"],
            reader_id="Persistence QA",
            review_round=1,
            atlasregion="QA region",
            initial_note="Manual point",
            session_id=session["session_id"],
        )
        targets = list_targets(self.db_path, self.case_id, "Persistence QA", 1)
        self.assertEqual(len(targets), 2)
        manual = next(item for item in targets if item["manual_id"] == manual_id.split(":", 1)[1])
        self.assertEqual(manual["origin"], "Manual")
        self.assertEqual(manual["manual_note"], "Manual point")
        logs = recent_case_log(self.db_path, self.case_id)
        event_types = {row["event_type"] for row in logs}
        self.assertIn("review_saved", event_types)
        self.assertIn("manual_microbleed_added", event_types)

    def test_other_readers_reports_are_visible_to_every_reader(self) -> None:
        first = start_new_session(self.db_path, "Reader One")
        first_target = list_targets(self.db_path, self.case_id, "Reader One", 1)[0]
        save_review(
            self.db_path,
            target_id=first_target["target_id"],
            case_id=self.case_id,
            reader_id="Reader One",
            review_round=1,
            verify=0,
            comment="First reader report",
            session_id=first["session_id"],
        )
        second_targets = list_targets(self.db_path, self.case_id, "Reader Two", 1)
        reports = second_targets[0]["reader_reports"]
        self.assertTrue(any(report["reader_id"] == "Reader One" for report in reports))
        self.assertTrue(any(report["comment"] == "First reader report" for report in reports))

    def test_ras_mapping_and_canonical_volume(self) -> None:
        case = get_case(self.db_path, self.case_id)
        target = list_targets(self.db_path, self.case_id, "RAS QA", 1)[0]
        volume = load_volume(case["qsm_path"])
        voxel = ras_to_voxel(volume.affine, target["ras"])
        self.assertTrue(voxel_in_bounds(voxel, volume.shape))
        self.assertEqual(volume.orientation, ("L", "P", "I"))
        np.testing.assert_allclose(voxel, [151.6666631, 64.3333267, 55.0], atol=0.05)
        np.testing.assert_allclose(voxel_to_ras(volume.affine, voxel), target["ras"], atol=1e-5)

        # The selected finding and a manual jump through the same RAS values
        # must resolve to the exact same linked slice indices.
        selected_indices = np.rint(voxel).astype(int)
        manual_indices = np.rint(ras_to_voxel(volume.affine, tuple(float(x) for x in target["ras"]))).astype(int)
        np.testing.assert_array_equal(selected_indices, manual_indices)

    def test_case_list_can_support_filters(self) -> None:
        cases = list_cases(self.db_path, "Fresh QA", 1)
        self.assertEqual(len(cases), self.report["case_count"])
        f101279 = next(item for item in cases if item["case_id"] == self.case_id)
        self.assertEqual(f101279["file_status"], "complete")
        self.assertEqual(f101279["reader_review_status"], "Unreviewed")

    def test_a_manual_finding_counts_towards_the_case_progress(self) -> None:
        """A case is not finished while one of its findings is unreviewed.

        The queue counted only the workbook rows, so adding a microbleed the
        sheet had missed left the case reading "2/2 reviewed", coloured as
        complete, and out of the "not yet reviewed by me" filter.
        """

        temp_dir = Path(tempfile.mkdtemp(prefix="microbleed_manual_progress_"))
        try:
            db_path = temp_dir / "review.sqlite"
            initialize_store(self.source, self.data_root, db_path)
            reader = "Progress Reader"
            start_new_session(db_path, reader)

            def case_row() -> dict:
                return next(
                    item for item in list_cases(db_path, reader, 1) if item["case_id"] == self.case_id
                )

            before = case_row()
            source_targets = list_targets(db_path, self.case_id, reader, 1)
            for target in source_targets:
                save_review(
                    db_path, target_id=str(target["target_id"]), case_id=self.case_id,
                    reader_id=reader, review_round=1, verify=1, comment=None,
                )
            finished = case_row()
            self.assertEqual(finished["reader_review_status"], "Reviewed")

            manual_target = add_manual_annotation(
                db_path, case_id=self.case_id, ras=(1.0, 2.0, 3.0), reader_id=reader,
                review_round=1, atlasregion="extra", initial_note="missed on the sheet",
            )
            with_manual = case_row()
            self.assertEqual(
                with_manual["finding_count"], before["finding_count"] + 1,
                "the manual finding is missing from the case's finding count",
            )
            self.assertEqual(
                with_manual["reader_review_status"], "In progress",
                "a case with an unreviewed manual finding must not read as complete",
            )
            # The source-only count keeps its meaning, so the source filters do
            # not start matching cases just because someone added a finding.
            self.assertEqual(with_manual["source_count"], before["source_count"])
            self.assertEqual(
                with_manual["source_unverified_count"], before["source_unverified_count"]
            )

            save_review(
                db_path, target_id=manual_target, case_id=self.case_id,
                reader_id=reader, review_round=1, verify=1, comment=None,
            )
            self.assertEqual(case_row()["reader_review_status"], "Reviewed")
        finally:
            import shutil

            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_a_thread_can_ask_sqlite_for_less_patience(self) -> None:
        """The background writer must not inherit the GUI's 30 second wait.

        Its writes are the ones nobody is waiting for, so a locked database
        should make them fail quickly instead of holding the thread long enough
        that shutting down has to consider killing it.
        """

        import review_store

        connection = connect(self.db_path)
        try:
            default = int(connection.execute("PRAGMA busy_timeout").fetchone()[0])
        finally:
            connection.close()
        self.assertEqual(default, review_store.DEFAULT_BUSY_TIMEOUT_MS)

        review_store.set_busy_timeout_ms(500)
        try:
            connection = connect(self.db_path)
            try:
                self.assertEqual(int(connection.execute("PRAGMA busy_timeout").fetchone()[0]), 500)
            finally:
                connection.close()
        finally:
            review_store.set_busy_timeout_ms(None)

        connection = connect(self.db_path)
        try:
            self.assertEqual(
                int(connection.execute("PRAGMA busy_timeout").fetchone()[0]),
                review_store.DEFAULT_BUSY_TIMEOUT_MS,
            )
        finally:
            connection.close()

    def test_uniform_slice_uses_a_non_degenerate_window(self) -> None:
        low, high = robust_window(np.zeros((32, 32), dtype=np.float32), fallback=np.linspace(0, 1, 128))
        self.assertLess(low, high)


class ExportMergeAndReimportTests(unittest.TestCase):
    """Getting results out, combining readers, and following the workbook."""

    @classmethod
    def setUpClass(cls) -> None:
        source = os.environ.get("TEST_SOURCE_XLSX")
        if not source:
            raise unittest.SkipTest("Set TEST_SOURCE_XLSX to a readable copy of the source workbook.")
        cls.source = Path(source)
        cls.data_root = Path(os.environ.get("TEST_DATA_ROOT", VIEWER_DIR.parent / "Data"))

    def setUp(self) -> None:
        self.temp_dir = Path(tempfile.mkdtemp(prefix="microbleed_export_test_"))
        self.db_path = self.temp_dir / "review.sqlite"
        initialize_store(self.source, self.data_root, self.db_path)
        self.case_id = _sample_case(self.db_path)

    def tearDown(self) -> None:
        import shutil

        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _first_target(self, db_path: Path, reader: str) -> str:
        targets = list_targets(db_path, self.case_id, reader, 1)
        return str(targets[0]["target_id"])

    def test_export_reports_agreement_between_readers(self) -> None:
        target_id = self._first_target(self.db_path, "A")
        start_new_session(self.db_path, "Reader A")
        start_new_session(self.db_path, "Reader B")
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=1, comment="clear focus",
        )
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader B",
            review_round=1, verify=0, comment="vessel",
        )
        out_path = self.temp_dir / "export.xlsx"
        result = export_reviews(self.db_path, out_path)
        self.assertTrue(out_path.exists())
        self.assertEqual(result["readers"], 2)
        self.assertEqual(result["reader_reports"], 2)
        self.assertEqual(result["disagreements"], 1)

        _readers, findings, _long = collect_export_rows(self.db_path)
        row = next(item for item in findings if item["target_id"] == target_id)
        self.assertEqual(row["Reader A · verify"], 1)
        self.assertEqual(row["Reader B · verify"], 0)
        self.assertEqual(row["agreement"], "disagreement")
        self.assertEqual(row["readers_with_verdict"], 2)
        untouched = next(item for item in findings if item["target_id"] != target_id)
        self.assertEqual(untouched["agreement"], "no verdict")

    def test_a_verdict_carries_how_sure_and_what_else_it_might_be(self) -> None:
        """"No" with a free-text reason cannot be counted; a category can."""

        target_id = self._first_target(self.db_path, "A")
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=0, comment="looks linear",
            certainty="definite", mimic="vessel",
        )
        targets = list_targets(self.db_path, self.case_id, "Reader A", 1)
        finding = next(item for item in targets if item["target_id"] == target_id)
        self.assertEqual(finding["reader_certainty"], "definite")
        self.assertEqual(finding["reader_mimic"], "vessel")

        _readers, findings, long_rows = collect_export_rows(self.db_path)
        row = next(item for item in findings if item["target_id"] == target_id)
        self.assertEqual(row["Reader A · certainty"], "definite")
        self.assertEqual(row["Reader A · mimic"], "vessel")
        report = next(item for item in long_rows if item["target_id"] == target_id)
        self.assertEqual(report["certainty"], "definite")
        self.assertEqual(report["mimic"], "vessel")

    def test_the_queue_knows_where_the_readers_disagree(self) -> None:
        """The source sheet's adjudication notes answer a different question.

        Which findings *this* study's readers have decided differently is the
        queue a second round works from, and it could not be asked for.
        """

        target_id = self._first_target(self.db_path, "A")
        def case_row(reader: str) -> dict:
            return next(
                item for item in list_cases(self.db_path, reader, 1)
                if item["case_id"] == self.case_id
            )

        self.assertEqual(case_row("Reader A")["disagreement_count"], 0)
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=1, comment=None,
        )
        self.assertEqual(case_row("Reader A")["disagreement_count"], 0, "one reader cannot disagree")
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader B",
            review_round=1, verify=1, comment=None,
        )
        self.assertEqual(case_row("Reader A")["disagreement_count"], 0, "they agreed")
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader B",
            review_round=1, verify=0, comment=None,
        )
        # Visible to either reader: it is a property of the finding.
        self.assertEqual(case_row("Reader A")["disagreement_count"], 1)
        self.assertEqual(case_row("Reader B")["disagreement_count"], 1)

    def test_two_readers_segmentations_are_compared(self) -> None:
        """Dice and centroid distance are what a reliability section reports.

        Neither can be worked out from the volumes alone: two masks of the
        same size can sit in different places.
        """

        from review_store import AGREEMENT_COLUMNS, collect_agreement_rows

        target_id = self._first_target(self.db_path, "A")
        for reader, offset, value in (("Reader A", 0, 1), ("Reader B", 1, 4)):
            start_new_session(self.db_path, reader)
            path = label_path(self.db_path, self.case_id, reader, 1)
            path.parent.mkdir(parents=True, exist_ok=True)
            data = np.zeros((10, 10, 10), dtype=np.uint16)
            # Two 3x3x3 cubes that overlap in two of three columns.
            data[2 + offset:5 + offset, 2:5, 2:5] = value
            import nibabel as nib

            nib.save(nib.Nifti1Image(data, np.eye(4)), str(path))
            save_roi(
                self.db_path, target_id=target_id, case_id=self.case_id, reader_id=reader,
                review_round=1, label_value=value, path=path, voxel_count=27,
                volume_mm3=27.0, generated_from="swi",
                centroid_ras=(3.0 + offset, 3.0, 3.0),
            )

        rows = collect_agreement_rows(self.db_path)
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual({row["reader_a"], row["reader_b"]}, {"Reader A", "Reader B"})
        # 18 shared voxels of 27 + 27.
        self.assertAlmostEqual(float(row["dice"]), 2 * 18 / 54, places=4)
        self.assertAlmostEqual(float(row["centroid_mm"]), 1.0, places=4)
        self.assertAlmostEqual(float(row["volume_ratio"]), 1.0, places=4)
        for column in ("dice", "centroid_mm", "volume_ratio"):
            self.assertIn(column, AGREEMENT_COLUMNS)

        out_path = self.temp_dir / "export.xlsx"
        result = export_reviews(self.db_path, out_path)
        self.assertEqual(result["segmentation_pairs"], 1)
        from openpyxl import load_workbook

        book = load_workbook(out_path)
        self.assertIn("segmentation_agreement", book.sheetnames)

    def test_a_finding_only_one_reader_segmented_is_not_compared(self) -> None:
        from review_store import collect_agreement_rows

        target_id = self._first_target(self.db_path, "A")
        save_roi(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Solo",
            review_round=1, label_value=1, path=self.temp_dir / "solo.nii.gz",
            voxel_count=20, volume_mm3=14.8, generated_from="swi",
        )
        self.assertEqual(collect_agreement_rows(self.db_path), [])

    def test_export_to_csv(self) -> None:
        target_id = self._first_target(self.db_path, "A")
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Solo",
            review_round=1, verify=1, comment=None,
        )
        out_path = self.temp_dir / "export.csv"
        result = export_reviews(self.db_path, out_path)
        text = out_path.read_text(encoding="utf-8-sig")
        self.assertIn("Solo · verify", text.splitlines()[0])
        with connect(self.db_path) as connection:
            expected = connection.execute(
                "SELECT COUNT(*) AS n FROM source_microbleeds"
            ).fetchone()["n"]
        self.assertEqual(result["findings"], expected)
        self.assertEqual(len(text.splitlines()) - 1, expected, "a row went missing")

    def test_merge_keeps_both_readers_and_renumbers_colliding_rounds(self) -> None:
        target_id = self._first_target(self.db_path, "A")
        # Two readers, each with their own database, both using round 1.
        first = self.temp_dir / "reader_a.sqlite"
        second = self.temp_dir / "reader_b.sqlite"
        for path, reader, verdict, comment in (
            (first, "Reader A", 1, "definitely"),
            (second, "Reader B", 0, "mimic"),
        ):
            initialize_store(self.source, self.data_root, path)
            start_new_session(path, reader)
            save_review(
                path, target_id=target_id, case_id=self.case_id, reader_id=reader,
                review_round=1, verify=verdict, comment=comment,
            )
        # And a same-named reader in both, which must not overwrite anything.
        start_new_session(first, "Shared Reader")
        save_review(
            first, target_id=target_id, case_id=self.case_id, reader_id="Shared Reader",
            review_round=1, verify=1, comment="from A",
        )
        start_new_session(second, "Shared Reader")
        save_review(
            second, target_id=target_id, case_id=self.case_id, reader_id="Shared Reader",
            review_round=1, verify=0, comment="from B",
        )

        merged = self.temp_dir / "merged.sqlite"
        initialize_store(self.source, self.data_root, merged)
        summary = merge_stores(merged, [first, second])
        self.assertEqual(summary["reviews_added"], 4)
        self.assertGreaterEqual(summary["rounds_renumbered"], 1)

        _readers, findings, long_rows = collect_export_rows(merged)
        row = next(item for item in findings if item["target_id"] == target_id)
        self.assertEqual(row["Reader A · verify"], 1)
        self.assertEqual(row["Reader B · verify"], 0)
        # Both of the same-named reader's rounds survived, in different rounds.
        shared = [item for item in long_rows if item["reader_id"] == "Shared Reader"]
        self.assertEqual(len(shared), 2)
        self.assertEqual({item["comment"] for item in shared}, {"from A", "from B"})
        self.assertEqual(len({item["review_round"] for item in shared}), 2)

    def test_merging_twice_does_not_duplicate(self) -> None:
        target_id = self._first_target(self.db_path, "A")
        source_db = self.temp_dir / "one.sqlite"
        initialize_store(self.source, self.data_root, source_db)
        start_new_session(source_db, "Reader A")
        save_review(
            source_db, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=1, comment="once",
        )
        merged = self.temp_dir / "merged.sqlite"
        initialize_store(self.source, self.data_root, merged)
        merge_stores(merged, [source_db])
        merge_stores(merged, [source_db])
        _readers, _findings, long_rows = collect_export_rows(merged)
        self.assertEqual(len(long_rows), 1)

    def _write_label_file(self, path: Path, value: int = 1) -> Path:
        """A small but real label NIfTI, so merging has a file to carry."""

        import nibabel as nib

        path.parent.mkdir(parents=True, exist_ok=True)
        data = np.zeros((6, 6, 6), dtype=np.uint16)
        data[2:4, 2:4, 2:4] = value
        nib.save(nib.Nifti1Image(data, np.eye(4)), str(path))
        return path

    def test_segmentations_survive_a_merge(self) -> None:
        """Per-reader databases are the recommended multi-reader workflow, so a
        merge that drops ``roi_labels`` loses every segmentation silently."""

        import nibabel as nib

        target_id = self._first_target(self.db_path, "A")
        source_db = self.temp_dir / "reader_a.sqlite"
        initialize_store(self.source, self.data_root, source_db)
        start_new_session(source_db, "Reader A")
        save_review(
            source_db, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=1, comment="segmented",
        )
        label_file = self._write_label_file(
            label_path(source_db, self.case_id, "Reader A", 1), value=3
        )
        save_roi(
            source_db, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, label_value=3, path=label_file, voxel_count=8,
            volume_mm3=8.0, generated_from="swi", centroid_ras=(1.5, 2.5, 3.5),
        )

        merged = self.temp_dir / "merged.sqlite"
        initialize_store(self.source, self.data_root, merged)
        merge_stores(merged, [source_db])

        rois = list_rois(merged, self.case_id, "Reader A", 1)
        self.assertIn(target_id, rois)
        row = rois[target_id]
        self.assertEqual(row["label_value"], 3)
        self.assertEqual(row["voxel_count"], 8)
        self.assertEqual(row["generated_from"], "swi")
        self.assertAlmostEqual(float(row["centroid_l"]), 1.5)

        # The mask file has to travel with the row, and the row has to point at
        # the copy: the source reader's folder may not exist on this machine.
        copied = label_path(merged, self.case_id, "Reader A", 1)
        self.assertTrue(copied.exists(), "the label file was not copied beside the merged store")
        self.assertEqual(Path(row["path"]), copied)
        self.assertEqual(int(np.asarray(nib.load(str(copied)).dataobj).max()), 3)

        segmentations = collect_segmentation_rows(merged)
        self.assertEqual(len(segmentations), 1)
        self.assertEqual(segmentations[0]["label_value"], 3)

    def test_a_re_merged_review_brings_its_corrected_position_with_it(self) -> None:
        """The update path used to refresh verify and comment but not ras_*."""

        target_id = self._first_target(self.db_path, "A")
        source_db = self.temp_dir / "reader_a.sqlite"
        initialize_store(self.source, self.data_root, source_db)
        start_new_session(source_db, "Reader A")
        save_review(
            source_db, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=1, comment="first look", corrected_ras=(10.0, 20.0, 30.0),
        )
        merged = self.temp_dir / "merged.sqlite"
        initialize_store(self.source, self.data_root, merged)
        merge_stores(merged, [source_db])

        # The reader looks again and moves the finding somewhere else.
        self._touch_review(
            source_db, target_id, "Reader A",
            verify=0, comment="a vessel after all", ras=(99.0, 98.0, 97.0),
        )
        merge_stores(merged, [source_db])

        connection = connect(merged)
        try:
            row = connection.execute(
                "SELECT verify, comment, ras_l, ras_p, ras_s FROM review_annotations "
                "WHERE target_id = ? AND reader_id = 'Reader A'",
                (target_id,),
            ).fetchone()
        finally:
            connection.close()
        self.assertEqual(row["verify"], 0)
        self.assertEqual(row["comment"], "a vessel after all")
        self.assertAlmostEqual(float(row["ras_l"]), 99.0)
        self.assertAlmostEqual(float(row["ras_p"]), 98.0)
        self.assertAlmostEqual(float(row["ras_s"]), 97.0)

    def _touch_review(self, db_path: Path, target_id: str, reader: str, *, verify, comment, ras) -> None:
        """Re-save a review with an ``updated_at`` the merge will treat as newer.

        ``utc_now`` has one-second resolution, so a test that saved twice in the
        same second would not exercise the update path at all.
        """

        save_review(
            db_path, target_id=target_id, case_id=self.case_id, reader_id=reader,
            review_round=1, verify=verify, comment=comment, corrected_ras=ras,
        )
        connection = connect(db_path)
        try:
            connection.execute(
                "UPDATE review_annotations SET updated_at = '2099-01-01T00:00:00+00:00' "
                "WHERE target_id = ? AND reader_id = ?",
                (target_id, reader),
            )
            connection.commit()
        finally:
            connection.close()

    def test_re_merging_does_not_duplicate_the_operation_log(self) -> None:
        """``merge_stores`` is documented as repeatable; the log has to obey."""

        target_id = self._first_target(self.db_path, "A")
        source_db = self.temp_dir / "reader_a.sqlite"
        initialize_store(self.source, self.data_root, source_db)
        start_new_session(source_db, "Reader A")
        save_review(
            source_db, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=1, comment="once",
        )
        merged = self.temp_dir / "merged.sqlite"
        initialize_store(self.source, self.data_root, merged)
        merge_stores(merged, [source_db])

        def log_count() -> int:
            connection = connect(merged)
            try:
                return int(connection.execute("SELECT COUNT(*) AS n FROM operation_log").fetchone()["n"])
            finally:
                connection.close()

        before = log_count()
        self.assertGreater(before, 0)
        merge_stores(merged, [source_db])
        self.assertEqual(log_count(), before)

        # New work in the source still has to arrive on a later merge.
        save_review(
            source_db, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=0, comment="changed my mind",
        )
        merge_stores(merged, [source_db])
        self.assertGreater(log_count(), before)

    def test_a_reader_can_remove_a_finding_they_added(self) -> None:
        """A finding added by hand is the one kind that can be a mistake.

        The workbook rows are given; a manual one is an observation this
        reader made, so undoing it has to be possible without editing the
        database by hand.
        """

        from review_store import delete_manual_annotation, manual_deletion_blockers

        start_new_session(self.db_path, "Reader A")
        target_id = add_manual_annotation(
            self.db_path, case_id=self.case_id, ras=(1.0, 2.0, 3.0), reader_id="Reader A",
            review_round=1, atlasregion="extra", initial_note="a mistake",
        )
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=1, comment="mine",
        )
        save_roi(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, label_value=9, path=self.temp_dir / "l.nii.gz",
            voxel_count=12, volume_mm3=8.9, generated_from="swi",
        )

        self.assertEqual(manual_deletion_blockers(self.db_path, target_id, "Reader A"), [])
        removed = delete_manual_annotation(
            self.db_path, target_id=target_id, reader_id="Reader A",
        )
        self.assertEqual(removed["reviews"], 1)
        self.assertEqual(removed["segmentations"], 1)

        remaining = list_targets(self.db_path, self.case_id, "Reader A", 1)
        self.assertNotIn(target_id, [item["target_id"] for item in remaining])
        connection = connect(self.db_path)
        try:
            for table in ("review_annotations", "roi_labels"):
                left = connection.execute(
                    f"SELECT COUNT(*) AS n FROM {table} WHERE target_id = ?",  # noqa: S608
                    (target_id,),
                ).fetchone()["n"]
                self.assertEqual(left, 0, f"{table} kept a row for a deleted finding")
        finally:
            connection.close()

    def test_a_source_finding_and_another_readers_work_are_not_deletable(self) -> None:
        """Deleting has to stop at the point where it destroys someone else's
        observation, or a shared datasheet is not safe to work in."""

        from review_store import delete_manual_annotation, manual_deletion_blockers

        source_id = self._first_target(self.db_path, "A")
        blockers = manual_deletion_blockers(self.db_path, source_id, "Reader A")
        self.assertTrue(blockers)
        self.assertIn("workbook", " ".join(blockers).lower())
        with self.assertRaises(ValueError):
            delete_manual_annotation(self.db_path, target_id=source_id, reader_id="Reader A")

        target_id = add_manual_annotation(
            self.db_path, case_id=self.case_id, ras=(1.0, 2.0, 3.0), reader_id="Reader A",
            review_round=1,
        )
        # Somebody else's finding.
        blockers = manual_deletion_blockers(self.db_path, target_id, "Reader B")
        self.assertTrue(any("Reader A" in reason for reason in blockers))

        # Their own, but another reader has already judged it.
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader B",
            review_round=1, verify=0, comment="not one",
        )
        blockers = manual_deletion_blockers(self.db_path, target_id, "Reader A")
        self.assertTrue(any("Reader B" in reason for reason in blockers))
        with self.assertRaises(ValueError):
            delete_manual_annotation(self.db_path, target_id=target_id, reader_id="Reader A")

    def test_a_finding_already_at_that_spot_is_found_before_adding_another(self) -> None:
        """Two distinct findings in one case never come within 2.03 mm here.

        Measured over all 1016 within-case pairs in this dataset: none closer
        than 2 mm, one at 2.03, three under 4.  So a point within a millimetre
        of an existing finding is that finding, and one within three is worth
        asking about.
        """

        from review_store import NEARBY_FINDING_MM, SAME_FINDING_MM, findings_near

        source = list_targets(self.db_path, self.case_id, "R", 1)[0]
        ras = tuple(float(value) for value in source["ras"])

        exact = findings_near(self.db_path, self.case_id, ras)
        self.assertTrue(exact)
        self.assertEqual(exact[0]["target_id"], source["target_id"])
        self.assertLess(exact[0]["distance_mm"], 1e-6)

        # Half a millimetre away is the same lesion.
        close = findings_near(self.db_path, self.case_id, (ras[0] + 0.5, ras[1], ras[2]))
        self.assertTrue(close and close[0]["distance_mm"] < SAME_FINDING_MM)

        # Two millimetres away is worth a question, not a refusal.
        nearby = findings_near(self.db_path, self.case_id, (ras[0] + 2.0, ras[1], ras[2]))
        self.assertTrue(nearby)
        self.assertGreater(nearby[0]["distance_mm"], SAME_FINDING_MM)
        self.assertLess(nearby[0]["distance_mm"], NEARBY_FINDING_MM)

        # Far away is nobody's business.
        self.assertEqual(findings_near(self.db_path, self.case_id, (ras[0] + 40.0, ras[1], ras[2])), [])
        # And a distinct case is never consulted.
        other_case = next(
            item["case_id"]
            for item in list_cases(self.db_path, "near QA", 1)
            if item["case_id"] != self.case_id
        )
        self.assertEqual(findings_near(self.db_path, other_case, ras), [])

    def test_merge_preserves_manual_annotations(self) -> None:
        source_db = self.temp_dir / "manual.sqlite"
        initialize_store(self.source, self.data_root, source_db)
        start_new_session(source_db, "Reader A")
        manual_target = add_manual_annotation(
            source_db, case_id=self.case_id, ras=(1.0, 2.0, 3.0), reader_id="Reader A",
            review_round=1, atlasregion="extra", initial_note="missed on the sheet",
        )
        merged = self.temp_dir / "merged.sqlite"
        initialize_store(self.source, self.data_root, merged)
        merge_stores(merged, [source_db])
        targets = list_targets(merged, self.case_id, "Reader A", 1)
        manual = [item for item in targets if item["target_id"] == manual_target]
        self.assertEqual(len(manual), 1)
        self.assertEqual(manual[0]["atlasregion"], "extra")

    def test_an_older_store_gains_the_correction_columns(self) -> None:
        """A database written before corrections existed must keep working."""

        legacy = self.temp_dir / "legacy.sqlite"
        with connect(legacy) as connection:
            connection.executescript(
                """
                CREATE TABLE review_annotations (
                    review_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    target_id TEXT NOT NULL,
                    case_id TEXT NOT NULL,
                    reader_id TEXT NOT NULL,
                    review_round INTEGER NOT NULL,
                    verify INTEGER,
                    comment TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(target_id, reader_id, review_round)
                );
                INSERT INTO review_annotations(
                    target_id, case_id, reader_id, review_round, verify, comment,
                    created_at, updated_at
                ) VALUES ('source:17', 'CASE0001', 'Old Reader', 1, 1, 'from before',
                          '2026-01-01T00:00:00+00:00', '2026-01-01T00:00:00+00:00');
                """
            )
            connection.commit()

        initialize_store(self.source, self.data_root, legacy)
        with connect(legacy) as connection:
            columns = {row["name"] for row in connection.execute("PRAGMA table_info(review_annotations)")}
        self.assertTrue({"ras_l", "ras_p", "ras_s"} <= columns)
        targets = list_targets(legacy, self.case_id, "Old Reader", 1)
        finding = next(item for item in targets if item["target_id"] == "source:17")
        # The old review survived, and reads as "accepts the source position".
        self.assertEqual(finding["reader_verify"], 1)
        self.assertEqual(finding["reader_comment"], "from before")
        self.assertIsNone(finding["reader_ras"])
        self.assertEqual(len(finding["position_variants"]), 1)

    def test_an_older_store_gains_the_log_origin_columns(self) -> None:
        """Opening a database that predates a schema addition must not fail.

        The origin columns arrived with an index over them.  Creating that
        index in the schema script rather than in the migration made every
        existing store unopenable: ``CREATE TABLE IF NOT EXISTS`` is a no-op
        for a table that is already there, so the index referred to columns
        the migration had not added yet.  A fresh database hid it entirely.
        """

        legacy = self.temp_dir / "legacy_log.sqlite"
        with connect(legacy) as connection:
            connection.executescript(
                """
                CREATE TABLE operation_log (
                    log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    session_id TEXT,
                    reader_id TEXT,
                    review_round INTEGER,
                    event_type TEXT NOT NULL,
                    case_id TEXT,
                    target_id TEXT,
                    details_json TEXT,
                    created_at TEXT NOT NULL
                );
                INSERT INTO operation_log(event_type, case_id, created_at)
                VALUES ('case_loaded', 'CASE0001', '2026-01-01T00:00:00+00:00');
                """
            )
            connection.commit()

        initialize_store(self.source, self.data_root, legacy)
        with connect(legacy) as connection:
            columns = {row["name"] for row in connection.execute("PRAGMA table_info(operation_log)")}
            kept = connection.execute("SELECT COUNT(*) AS n FROM operation_log").fetchone()["n"]
            indexes = {row["name"] for row in connection.execute("PRAGMA index_list(operation_log)")}
        self.assertTrue({"origin_store_id", "origin_log_id"} <= columns)
        self.assertIn("idx_log_origin", indexes)
        self.assertGreaterEqual(kept, 1, "the existing log must survive the migration")

        # And the store is fully usable afterwards, including another merge.
        merged = self.temp_dir / "merged_legacy.sqlite"
        initialize_store(self.source, self.data_root, merged)
        merge_stores(merged, [legacy])
        merge_stores(merged, [legacy])
        with connect(merged) as connection:
            copies = connection.execute(
                "SELECT COUNT(*) AS n FROM operation_log WHERE event_type = 'case_loaded'"
            ).fetchone()["n"]
        self.assertEqual(copies, 1, "the pre-existing log row was merged twice")

    def test_corrections_survive_a_merge_and_reach_the_export(self) -> None:
        target_id = self._first_target(self.db_path, "A")
        source = list_targets(self.db_path, self.case_id, "A", 1)[0]["source_ras"]
        moved = tuple(value + 3.0 for value in source)
        reader_db = self.temp_dir / "reader.sqlite"
        initialize_store(self.source, self.data_root, reader_db)
        start_new_session(reader_db, "Reader A")
        save_review(
            reader_db, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=1, comment="moved it", corrected_ras=moved,
        )
        merged = self.temp_dir / "merged.sqlite"
        initialize_store(self.source, self.data_root, merged)
        merge_stores(merged, [reader_db])

        finding = next(
            item for item in list_targets(merged, self.case_id, "Reader A", 1)
            if item["target_id"] == target_id
        )
        for value, expected in zip(finding["reader_ras"], moved):
            self.assertAlmostEqual(value, expected, places=6)

        out_path = self.temp_dir / "with_corrections.xlsx"
        export_reviews(merged, out_path)
        _readers, findings, long_rows = collect_export_rows(merged)
        row = next(item for item in findings if item["target_id"] == target_id)
        self.assertAlmostEqual(row["Reader A · moved_mm"], 27.0 ** 0.5, places=3)
        self.assertEqual(row["readers_who_moved_it"], 1)
        report = next(item for item in long_rows if item["reader_id"] == "Reader A")
        self.assertAlmostEqual(report["ras_l"], moved[0], places=6)

    def test_reimport_adds_new_rows_and_keeps_reviews(self) -> None:
        from openpyxl import load_workbook

        target_id = self._first_target(self.db_path, "A")
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, verify=1, comment="keep me",
        )
        edited = self.temp_dir / "edited.xlsx"
        workbook = load_workbook(self.source)
        sheet = workbook["MCH-microhemorrage"]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        new_row = [None] * len(headers)
        for name, value in (
            ("subjectid", self.case_id),
            ("RAS L-R L", 1.5),
            ("RAS P-A A", 2.5),
            ("RAS I-S S", 3.5),
            ("verify (yes=1)", 1),
            ("atlasregions", "added_later"),
        ):
            new_row[headers.index(name)] = value
        sheet.append(new_row)
        # Also edit the source comment of the finding that has been reviewed;
        # its target id carries the Excel row it came from.
        reviewed_row = int(target_id.split(":")[1])
        comment_column = headers.index("comments") + 1
        sheet.cell(row=reviewed_row, column=comment_column, value="revised source note")
        workbook.save(edited)

        result = reimport_source(self.db_path, edited)
        self.assertEqual(result["added"], 1)
        self.assertGreaterEqual(result["updated"], 1)
        self.assertEqual(result["removed_from_workbook"], 0)

        targets = list_targets(self.db_path, self.case_id, "Reader A", 1)
        self.assertIn("added_later", [item["atlasregion"] for item in targets])
        reviewed = next(item for item in targets if item["target_id"] == target_id)
        self.assertEqual(reviewed["reader_verify"], 1)
        self.assertEqual(reviewed["reader_comment"], "keep me")
        self.assertEqual(reviewed["source_comments"], "revised source note")

    def test_reimport_reports_rows_that_left_the_workbook(self) -> None:
        from openpyxl import load_workbook

        shortened = self.temp_dir / "short.xlsx"
        workbook = load_workbook(self.source)
        sheet = workbook["MCH-microhemorrage"]
        sheet.delete_rows(sheet.max_row)
        workbook.save(shortened)
        with connect(self.db_path) as connection:
            before = connection.execute(
                "SELECT COUNT(*) AS n FROM source_microbleeds"
            ).fetchone()["n"]
        result = reimport_source(self.db_path, shortened)
        self.assertEqual(result["removed_from_workbook"], 1)
        self.assertEqual(result["added"], 0)
        # Nothing was deleted from the store: a row that vanished from the
        # workbook may already have been reviewed.
        with connect(self.db_path) as connection:
            remaining = connection.execute(
                "SELECT COUNT(*) AS n FROM source_microbleeds"
            ).fetchone()["n"]
        self.assertEqual(remaining, before)


class RegionGrowingTests(unittest.TestCase):
    """The assisted segmentation has to behave on synthetic ground truth."""

    def _phantom(self, radius_mm=2.0, spacing=(0.859, 0.859, 1.0), dark=True):
        shape = (40, 40, 40)
        data = np.full(shape, 100.0, dtype=np.float32)
        centre = np.array([20, 20, 20])
        grids = np.ogrid[tuple(slice(0, size) for size in shape)]
        distance_sq = sum(
            ((grid - centre[axis]) * spacing[axis]) ** 2 for axis, grid in enumerate(grids)
        )
        sphere = distance_sq <= radius_mm**2
        data[sphere] = 10.0 if dark else 400.0
        # A little noise, so the threshold is not trivially exact.
        data += np.random.default_rng(0).normal(0, 1.5, shape).astype(np.float32)
        return data, centre, sphere, spacing

    def test_growing_recovers_a_synthetic_lesion(self) -> None:
        for dark in (True, False):
            data, centre, sphere, spacing = self._phantom(dark=dark)
            mask = grow_lesion(data, centre, spacing, dark=dark, sensitivity=1.5, radius_mm=6.0)
            with self.subTest(dark=dark):
                overlap = np.logical_and(mask, sphere).sum()
                dice = 2 * overlap / (mask.sum() + sphere.sum())
                self.assertGreater(dice, 0.85, f"dice {dice:.2f}")

    def test_growth_cannot_escape_the_radius(self) -> None:
        # A long dark bar through the seed, like a vessel.
        shape = (40, 40, 40)
        spacing = (0.859, 0.859, 1.0)
        data = np.full(shape, 100.0, dtype=np.float32)
        data[19:22, 19:22, :] = 10.0
        mask = grow_lesion(data, (20, 20, 20), spacing, dark=True, sensitivity=1.5, radius_mm=4.0)
        extent = np.argwhere(mask)
        span_mm = (extent.max(0) - extent.min(0) + 1) * np.array(spacing)
        self.assertLessEqual(span_mm[2], 2 * 4.0 + 1.0, f"grew {span_mm[2]:.1f} mm along the bar")

    def test_the_seed_is_always_included(self) -> None:
        data = np.full((20, 20, 20), 50.0, dtype=np.float32)
        mask = grow_lesion(data, (10, 10, 10), (1.0, 1.0, 1.0), dark=True)
        # A flat volume has no lesion, but the reader pointed at this voxel.
        self.assertFalse(mask.any())
        data[10, 10, 10] = 0.0
        mask = grow_lesion(data, (10, 10, 10), (1.0, 1.0, 1.0), dark=True)
        self.assertTrue(mask[10, 10, 10])

    def test_a_seed_outside_the_volume_is_harmless(self) -> None:
        data = np.zeros((10, 10, 10), dtype=np.float32)
        mask = grow_lesion(data, (99, 99, 99), (1.0, 1.0, 1.0), dark=True)
        self.assertFalse(mask.any())

    def _brain_like_phantom(self, lesion_mm=2.5, spacing=(0.859, 0.859, 1.0)):
        """A lesion with other anatomy nearby, which is what a cube picks up.

        A uniform background hides the whole problem: its spread comes from the
        noise and barely moves as the cube grows.  Real tissue has structure a
        centimetre away -- another vessel, a ventricle edge -- so the spread of
        a cube depends on how big the cube is, and the cube is sized by the
        growth cap.
        """

        shape = (48, 48, 48)
        centre = np.array([24, 24, 24])
        data = np.full(shape, 100.0, dtype=np.float32)
        grids = np.ogrid[tuple(slice(0, size) for size in shape)]
        distance_sq = sum(
            ((grid - centre[axis]) * spacing[axis]) ** 2 for axis, grid in enumerate(grids)
        )
        sphere = distance_sq <= lesion_mm**2
        data[sphere] = 40.0
        # Structure that a 4 mm cube misses and a 12 mm cube contains.
        data[:, :, :14] = 190.0          # a bright band, like a different tissue
        data[36:, :, :] = 25.0           # a dark structure off to one side
        data += np.random.default_rng(1).normal(0, 1.5, shape).astype(np.float32)
        return data, centre, sphere, spacing

    def test_the_growth_cap_does_not_decide_the_answer(self) -> None:
        """``radius_mm`` is a safety cap, not a segmentation parameter.

        The threshold came from the standard deviation of a cube sized by that
        cap -- and the cube holds whatever anatomy is nearby, so enlarging the
        cap widened the spread, tightened the threshold, and could shrink a
        real 20 mm3 lesion to the single seed voxel.  Measured on this dataset
        the answer moved by up to a factor of 44 as the cap changed.
        """

        data, centre, _sphere, spacing = self._brain_like_phantom()
        volumes = []
        for radius_mm in (4.0, 6.0, 8.0, 10.0, 12.0):
            mask = grow_lesion(
                data, centre, spacing, dark=True, sensitivity=2.0, radius_mm=radius_mm
            )
            volumes.append(float(mask.sum()) * float(np.prod(spacing)))
        self.assertGreater(min(volumes), 2.0, "the lesion collapsed at some cap")
        self.assertLess(
            max(volumes) / min(volumes),
            1.2,
            f"the cap changed the measured volume: {[round(v, 1) for v in volumes]}",
        )

    def test_a_faint_lesion_still_grows_past_its_seed(self) -> None:
        """A spread inflated by nearby anatomy rejected the lesion itself.

        Two real findings in this dataset returned exactly one voxel -- the
        reader pressed Generate and got nothing, with no explanation.
        """

        data, centre, sphere, spacing = self._brain_like_phantom(lesion_mm=2.0)
        # A weak focus: clearly darker than its surroundings, but not by much.
        data[sphere] = 78.0
        mask = grow_lesion(data, centre, spacing, dark=True, sensitivity=2.0, radius_mm=8.0)
        self.assertGreater(mask.sum(), 8, "a faint lesion collapsed to its seed")

    def test_a_hole_inside_the_lesion_is_filled(self) -> None:
        data, centre, sphere, spacing = self._phantom(radius_mm=3.0)
        # A bright voxel in the middle of a dark lesion: noise, not background.
        data[centre[0] + 1, centre[1], centre[2]] = 100.0
        mask = grow_lesion(data, centre, spacing, dark=True, sensitivity=1.5, radius_mm=6.0)
        self.assertTrue(
            mask[centre[0] + 1, centre[1], centre[2]],
            "an enclosed voxel was left out of the lesion",
        )

    def test_a_vessel_leak_is_reported_not_hidden(self) -> None:
        """A mask that ran down a vessel has to be flagged, not quietly kept.

        Shrinking it would be guessing.  The reader needs to know the automatic
        result is unreliable here so they can redraw it.
        """

        from imaging import segment_lesion

        shape = (40, 40, 40)
        spacing = (0.859, 0.859, 1.0)
        data = np.full(shape, 100.0, dtype=np.float32)
        data[19:22, 19:22, :] = 10.0  # a bar through the seed, like a vessel
        _mask, vessel = segment_lesion(
            data, (20, 20, 20), spacing, dark=True, sensitivity=1.5, radius_mm=8.0
        )
        self.assertTrue(vessel["suspect"], "an obvious vessel was not flagged")
        self.assertTrue(vessel["reached_cap"], "growth ran into the cap and did not say so")
        self.assertGreater(vessel["longest_mm"], 10.0)

        # A lesion that stops on its own, well inside the cap, is not flagged.
        data, centre, _sphere, spacing = self._brain_like_phantom(lesion_mm=2.5)
        _round_mask, lesion = segment_lesion(
            data, centre, spacing, dark=True, sensitivity=2.0, radius_mm=8.0
        )
        self.assertFalse(lesion["suspect"], f"a contained lesion was flagged: {lesion}")
        self.assertFalse(lesion["reached_cap"])
        self.assertLess(lesion["longest_mm"], 10.0)

    def test_snapping_finds_the_middle_of_the_focus(self) -> None:
        """A reader clicking a 3 mm lesion cannot click its centre by eye.

        The coordinate they record is the one analysis uses, and half a
        millimetre of hand tremor is a fifth of a small microbleed.  Snapping
        to the darkest voxel nearby lets the data place the point.
        """

        from imaging import snap_to_extremum

        data, centre, _sphere, spacing = self._brain_like_phantom(lesion_mm=2.5)
        # Click one voxel off the centre in every direction, as a hand does.
        clicked = np.asarray(centre) + np.array([2, -2, 1])
        snapped = snap_to_extremum(data, clicked, spacing, dark=True, radius_mm=3.0)
        moved = np.linalg.norm((np.asarray(snapped) - np.asarray(centre)) * np.asarray(spacing))
        self.assertLess(moved, 1.5, f"snapped to {snapped}, wanted {centre}")

        # Bright lesions on QSM snap the other way.
        bright = 200.0 - data
        snapped = snap_to_extremum(bright, clicked, spacing, dark=False, radius_mm=3.0)
        moved = np.linalg.norm((np.asarray(snapped) - np.asarray(centre)) * np.asarray(spacing))
        self.assertLess(moved, 1.5)

    def test_snapping_stays_put_when_there_is_nothing_to_snap_to(self) -> None:
        """Flat tissue has no focus, and inventing one would be worse."""

        from imaging import snap_to_extremum

        flat = np.full((20, 20, 20), 100.0, dtype=np.float32)
        start = (10, 10, 10)
        np.testing.assert_array_equal(
            snap_to_extremum(flat, start, (1.0, 1.0, 1.0), dark=True, radius_mm=3.0), start
        )
        # And a point outside the volume is returned untouched, not clamped.
        outside = (99, 99, 99)
        np.testing.assert_array_equal(
            snap_to_extremum(flat, outside, (1.0, 1.0, 1.0), dark=True), outside
        )

    def test_the_shape_report_describes_an_empty_mask_without_failing(self) -> None:
        from imaging import lesion_shape

        empty = lesion_shape(np.zeros((8, 8, 8), dtype=bool), (1.0, 1.0, 1.0))
        self.assertEqual(empty["voxel_count"], 0)
        self.assertEqual(empty["volume_mm3"], 0.0)
        self.assertFalse(empty["suspect"])


class SegmentationStoreTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        source = os.environ.get("TEST_SOURCE_XLSX")
        if not source:
            raise unittest.SkipTest("Set TEST_SOURCE_XLSX to a readable copy of the source workbook.")
        cls.source = Path(source)
        cls.data_root = Path(os.environ.get("TEST_DATA_ROOT", VIEWER_DIR.parent / "Data"))

    def setUp(self) -> None:
        self.temp_dir = Path(tempfile.mkdtemp(prefix="microbleed_roi_test_"))
        self.db_path = self.temp_dir / "review.sqlite"
        initialize_store(self.source, self.data_root, self.db_path)
        self.case_id = _sample_case(self.db_path)

    def tearDown(self) -> None:
        import shutil

        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_roi_rows_round_trip_and_clear(self) -> None:
        target_id = list_targets(self.db_path, self.case_id, "A", 1)[0]["target_id"]
        path = label_path(self.db_path, self.case_id, "Reader A", 1)
        save_roi(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, label_value=1, path=path, voxel_count=34,
            volume_mm3=25.1, generated_from="swi",
        )
        rows = list_rois(self.db_path, self.case_id, "Reader A", 1)
        self.assertEqual(rows[target_id]["voxel_count"], 34)
        self.assertEqual(rows[target_id]["generated_from"], "swi")
        # It reaches the finding list too.
        finding = next(
            item for item in list_targets(self.db_path, self.case_id, "Reader A", 1)
            if item["target_id"] == target_id
        )
        self.assertAlmostEqual(finding["roi"]["volume_mm3"], 25.1, places=6)

        # Saving an empty mask removes the row rather than storing a zero.
        save_roi(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Reader A",
            review_round=1, label_value=1, path=path, voxel_count=0,
            volume_mm3=0.0, generated_from="swi",
        )
        self.assertEqual(list_rois(self.db_path, self.case_id, "Reader A", 1), {})

    def test_a_segmentation_records_how_it_was_made(self) -> None:
        """Volumetry is not reproducible without the settings behind it.

        ``generated_from`` said which sequence, but not which threshold, which
        cap, or whether a human drew it -- so two numbers in the same column
        could come from different procedures with nothing to say so.
        """

        target_id = list_targets(self.db_path, self.case_id, "R", 1)[0]["target_id"]
        save_roi(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="R",
            review_round=1, label_value=1, path=self.temp_dir / "labels.nii.gz",
            voxel_count=30, volume_mm3=22.1, generated_from="swi",
            centroid_ras=(1.0, 2.0, 3.0), method="grow+brush",
            sensitivity=3.0, radius_mm=6.0,
        )
        stored = list_rois(self.db_path, self.case_id, "R", 1)[target_id]
        self.assertEqual(stored["method"], "grow+brush")
        self.assertAlmostEqual(float(stored["sensitivity"]), 3.0)
        self.assertAlmostEqual(float(stored["radius_mm"]), 6.0)

        rows = collect_segmentation_rows(self.db_path)
        row = next(item for item in rows if item["target_id"] == target_id)
        self.assertEqual(row["method"], "grow+brush")
        self.assertAlmostEqual(float(row["sensitivity"]), 3.0)
        self.assertAlmostEqual(float(row["radius_mm"]), 6.0)
        for column in ("method", "sensitivity", "radius_mm"):
            self.assertIn(column, SEGMENTATION_COLUMNS)

        # A mask drawn entirely by hand has no threshold to report.
        other = list_targets(self.db_path, self.case_id, "R", 1)[0]["target_id"]
        save_roi(
            self.db_path, target_id=other, case_id=self.case_id, reader_id="R",
            review_round=1, label_value=2, path=self.temp_dir / "labels.nii.gz",
            voxel_count=12, volume_mm3=8.9, generated_from=None, method="brush",
        )
        hand = list_rois(self.db_path, self.case_id, "R", 1)[other]
        self.assertEqual(hand["method"], "brush")
        self.assertIsNone(hand["sensitivity"])

    def test_an_older_store_gains_the_segmentation_method_columns(self) -> None:
        legacy = self.temp_dir / "legacy_roi.sqlite"
        with connect(legacy) as connection:
            connection.executescript(
                """
                CREATE TABLE roi_labels (
                    roi_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    target_id TEXT NOT NULL,
                    case_id TEXT NOT NULL,
                    reader_id TEXT NOT NULL,
                    review_round INTEGER NOT NULL,
                    label_value INTEGER NOT NULL,
                    path TEXT NOT NULL,
                    voxel_count INTEGER NOT NULL,
                    volume_mm3 REAL NOT NULL,
                    generated_from TEXT,
                    updated_at TEXT NOT NULL,
                    UNIQUE(target_id, reader_id, review_round)
                );
                INSERT INTO roi_labels(target_id, case_id, reader_id, review_round,
                    label_value, path, voxel_count, volume_mm3, generated_from, updated_at)
                VALUES ('source:17', 'CASE0001', 'Old', 1, 1, 'old.nii.gz', 20, 14.8, 'swi',
                        '2026-01-01T00:00:00+00:00');
                """
            )
            connection.commit()

        initialize_store(self.source, self.data_root, legacy)
        # The same invented case the legacy row above was written under.
        stored = list_rois(legacy, "CASE0001", "Old", 1)["source:17"]
        self.assertEqual(stored["voxel_count"], 20)
        self.assertIsNone(stored["method"], "an old row cannot claim a method it never recorded")
        self.assertIsNone(stored["sensitivity"])

    def test_the_export_says_which_label_value_is_which_finding(self) -> None:
        """A label file holds a whole case; the sheet has to decode it.

        Two findings share one file under different integers, and they do not
        share a verdict: a reader can accept one and reject the other. Without
        the label value in the results table, nobody can tell which voxels the
        rejected verdict refers to.
        """

        case_id = next(
            (item["case_id"] for item in list_cases(self.db_path, "A", 1)
             if len(list_targets(self.db_path, item["case_id"], "A", 1)) >= 2),
            None,
        )
        if case_id is None:
            self.skipTest("No case in this data root has two findings.")
        targets = list_targets(self.db_path, case_id, "A", 1)
        first, second = targets[0]["target_id"], targets[1]["target_id"]
        path = label_path(self.db_path, case_id, "Reader A", 1)
        for target_id, value, verdict, volume in (
            (first, 1, 1, 30.0),
            (second, 2, 0, 12.5),
        ):
            save_review(
                self.db_path, target_id=target_id, case_id=case_id,
                reader_id="Reader A", review_round=1, verify=verdict, comment=None,
            )
            save_roi(
                self.db_path, target_id=target_id, case_id=case_id,
                reader_id="Reader A", review_round=1, label_value=value, path=path,
                voxel_count=int(volume), volume_mm3=volume, generated_from="swi",
                centroid_ras=(1.0 * value, 2.0 * value, 3.0 * value),
            )

        rows = {row["target_id"]: row for row in collect_segmentation_rows(self.db_path)}
        self.assertEqual(rows[first]["label_value"], 1)
        self.assertEqual(rows[second]["label_value"], 2)
        # Same file, opposite verdicts.
        self.assertEqual(rows[first]["label_file"], rows[second]["label_file"])
        self.assertEqual(rows[first]["verify"], 1)
        self.assertEqual(rows[second]["verify"], 0)
        # And where the voxels are, not only how many.
        self.assertAlmostEqual(rows[second]["centroid_l"], 2.0, places=6)
        self.assertAlmostEqual(rows[second]["centroid_s"], 6.0, places=6)

        out_path = self.temp_dir / "export.xlsx"
        result = export_reviews(self.db_path, out_path)
        self.assertEqual(result["segmentations"], 2)
        from openpyxl import load_workbook

        book = load_workbook(out_path)
        self.assertIn("segmentations", book.sheetnames)
        header = [cell.value for cell in book["segmentations"][1]]
        for column in ("label_value", "label_file", "verify", "centroid_l", "final_l"):
            self.assertIn(column, header)
        wide = [cell.value for cell in book["findings"][1]]
        self.assertIn("Reader A \u00b7 roi_label", wide)
        self.assertIn("Reader A \u00b7 roi_file", wide)

    def test_the_export_carries_the_coordinate_each_reader_stands_behind(self) -> None:
        """``moved_mm`` says how far, not where. Analysis needs the where."""

        target_id = list_targets(self.db_path, self.case_id, "A", 1)[0]["target_id"]
        source = next(
            item for item in list_targets(self.db_path, self.case_id, "A", 1)
            if item["target_id"] == target_id
        )["ras"]
        corrected = (source[0] + 1.5, source[1], source[2])
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Mover",
            review_round=1, verify=1, comment=None, corrected_ras=corrected,
        )
        save_review(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Keeper",
            review_round=1, verify=1, comment=None,
        )
        _readers, findings, _long = collect_export_rows(self.db_path)
        row = next(item for item in findings if item["target_id"] == target_id)
        # The reader who moved it reports their own coordinate...
        self.assertAlmostEqual(row["Mover \u00b7 final_l"], corrected[0], places=6)
        self.assertAlmostEqual(row["Mover \u00b7 moved_mm"], 1.5, places=3)
        # ...and the one who accepted it reports the source coordinate, not a blank.
        self.assertAlmostEqual(row["Keeper \u00b7 final_l"], source[0], places=6)
        self.assertIsNone(row["Keeper \u00b7 moved_mm"])

    def test_the_csv_export_keeps_the_label_map_beside_it(self) -> None:
        target_id = list_targets(self.db_path, self.case_id, "A", 1)[0]["target_id"]
        save_roi(
            self.db_path, target_id=target_id, case_id=self.case_id, reader_id="Solo",
            review_round=1, label_value=7, path=label_path(self.db_path, self.case_id, "Solo", 1),
            voxel_count=10, volume_mm3=8.0, generated_from="qsm",
        )
        out_path = self.temp_dir / "export.csv"
        export_reviews(self.db_path, out_path)
        sidecar = out_path.with_name("export_segmentations.csv")
        self.assertTrue(sidecar.exists())
        text = sidecar.read_text(encoding="utf-8-sig")
        self.assertIn("label_value", text.splitlines()[0])
        self.assertIn(",7,", text.replace("\r\n", "\n").splitlines()[1])

    def test_label_files_are_per_reader_and_round(self) -> None:
        first = label_path(self.db_path, self.case_id, "Reader A", 1)
        second = label_path(self.db_path, self.case_id, "Reader B", 1)
        third = label_path(self.db_path, self.case_id, "Reader A", 2)
        self.assertNotEqual(first, second)
        self.assertNotEqual(first, third)
        self.assertEqual(first.parent.parent.name, "labels")
        # A reader name with characters a filesystem dislikes still works.
        awkward = label_path(self.db_path, "F1", "Dr. X / Y", 1)
        self.assertNotIn("/", awkward.parent.name)

    def test_a_saved_label_is_voxel_aligned_with_the_original_file(self) -> None:
        """The mask must drop straight onto the image in any other viewer.

        The viewer works in a display orientation, so writing the array as it
        sits in memory would give the label a flipped affine relative to the
        NIfTI it belongs to: right in world space, but not voxel-for-voxel
        aligned, which is what every other tool assumes when it loads an image
        and a label together.
        """

        import nibabel as nib

        case = get_case(self.db_path, self.case_id)
        original = nib.load(case["swi_path"])
        for preset in ORIENTATION_PRESETS:
            reference = load_volume(case["swi_path"], preset_axcodes(preset))
            labels = np.zeros(reference.shape, dtype=np.uint16)
            labels[100:104, 60:64, 50:52] = 3
            path = self.temp_dir / "labels" / f"CASE0001_{preset}.nii.gz"
            save_label_volume(path, labels, reference)
            stored = nib.load(str(path))
            with self.subTest(preset=preset):
                self.assertEqual(stored.shape, original.shape)
                np.testing.assert_allclose(stored.affine, original.affine, atol=1e-6)
                self.assertEqual(
                    nib.aff2axcodes(stored.affine), nib.aff2axcodes(original.affine)
                )
                # Both space codes are set: some tools read sform, some qform,
                # and an unset qform is a classic "mask in the wrong place".
                self.assertEqual(int(stored.header["sform_code"]), int(original.header["sform_code"]))
                self.assertEqual(int(stored.header["qform_code"]), int(original.header["qform_code"]))
                self.assertEqual(stored.get_data_dtype(), np.uint16)

    def test_a_label_marks_the_same_world_point_it_was_drawn_on(self) -> None:
        import nibabel as nib

        case = get_case(self.db_path, self.case_id)
        target = list_targets(self.db_path, self.case_id, "QA", 1)[0]
        ras = tuple(float(value) for value in target["source_ras"])
        for preset in ORIENTATION_PRESETS:
            reference = load_volume(case["swi_path"], preset_axcodes(preset))
            voxel = np.rint(ras_to_voxel(reference.affine, ras)).astype(int)
            labels = np.zeros(reference.shape, dtype=np.uint16)
            labels[tuple(voxel)] = 5
            path = self.temp_dir / "labels" / f"point_{preset}.nii.gz"
            save_label_volume(path, labels, reference)
            stored = nib.load(str(path))
            marked = np.argwhere(np.asarray(stored.dataobj) == 5)
            with self.subTest(preset=preset):
                self.assertEqual(len(marked), 1)
                world = nib.affines.apply_affine(stored.affine, marked[0])
                # Round trip through the file lands back on the finding.
                self.assertLess(float(np.linalg.norm(world - np.asarray(ras))), 1.1)

    def test_the_label_file_does_not_depend_on_the_display_preset(self) -> None:
        import nibabel as nib

        case = get_case(self.db_path, self.case_id)
        written = {}
        for preset in ORIENTATION_PRESETS:
            reference = load_volume(case["swi_path"], preset_axcodes(preset))
            voxel = np.rint(
                ras_to_voxel(reference.affine, (-20.34, 45.15, 32.74))
            ).astype(int)
            labels = np.zeros(reference.shape, dtype=np.uint16)
            labels[tuple(voxel)] = 1
            path = self.temp_dir / f"preset_{preset}.nii.gz"
            save_label_volume(path, labels, reference)
            written[preset] = np.asarray(nib.load(str(path)).dataobj)
        presets = list(written)
        np.testing.assert_array_equal(written[presets[0]], written[presets[1]])


class OrientationPresetTests(unittest.TestCase):
    """The RAS contract must hold identically under every display preset.

    A display preset only mirrors what is drawn.  The physical meaning of a
    voxel is carried by the reoriented affine, so a finding has to land on the
    same tissue whichever preset is active.
    """

    @classmethod
    def setUpClass(cls) -> None:
        source = os.environ.get("TEST_SOURCE_XLSX")
        if not source:
            raise unittest.SkipTest("Set TEST_SOURCE_XLSX to a readable copy of the source workbook.")
        data_root = Path(os.environ.get("TEST_DATA_ROOT", VIEWER_DIR.parent / "Data"))
        matches = sorted(data_root.glob("*/*_GRE4_SWI_AffineRestored.nii.gz"))
        if not matches:
            raise unittest.SkipTest("No AffineRestored SWI volume available in the data root.")
        cls.path = matches[0]
        cls.volumes = {
            preset: load_volume(cls.path, preset_axcodes(preset)) for preset in ORIENTATION_PRESETS
        }
        # Sample physical points from the middle of the first volume so every
        # preset is asked about the same anatomy.
        reference = cls.volumes["radiological"]
        centre = np.asarray(reference.shape, dtype=float) / 2.0
        offsets = [(0, 0, 0), (12, -7, 5), (-20, 9, -11), (30, 25, 18), (-33, -21, 14)]
        cls.points = [
            tuple(float(x) for x in voxel_to_ras(reference.affine, centre + np.asarray(offset)))
            for offset in offsets
        ]

    def test_presets_are_distinct_left_right_conventions(self) -> None:
        codes = {preset: preset_axcodes(preset) for preset in ORIENTATION_PRESETS}
        self.assertEqual(codes["radiological"], ("L", "P", "I"))
        self.assertEqual(codes["neurological"], ("R", "P", "I"))
        for preset, axcodes in codes.items():
            with self.subTest(preset=preset):
                # Only the left-right axis may differ, so slice axes and plane
                # extraction stay valid for every preset.
                self.assertEqual(axcodes[1:], ("P", "I"))
                self.assertEqual(self.volumes[preset].orientation, axcodes)

    def test_ras_round_trips_in_every_preset(self) -> None:
        for preset, volume in self.volumes.items():
            for point in self.points:
                with self.subTest(preset=preset, point=point):
                    voxel = ras_to_voxel(volume.affine, point)
                    back = voxel_to_ras(volume.affine, voxel)
                    np.testing.assert_allclose(back, np.asarray(point), atol=1e-6)

    def test_the_same_ras_reaches_the_same_tissue_in_every_preset(self) -> None:
        for point in self.points:
            values = {}
            for preset, volume in self.volumes.items():
                voxel = ras_to_voxel(volume.affine, point)
                self.assertTrue(voxel_in_bounds(voxel, volume.shape), f"{preset} {point}")
                index = tuple(int(x) for x in clamp_voxel(voxel, volume.shape))
                values[preset] = float(volume.data[index])
            with self.subTest(point=point):
                # Identical voxel values mean the presets resolve one physical
                # point to one piece of anatomy, mirrored only on screen.
                self.assertEqual(len(set(values.values())), 1, values)

    def test_left_right_voxel_index_mirrors_between_presets(self) -> None:
        radiological = self.volumes["radiological"]
        neurological = self.volumes["neurological"]
        self.assertEqual(radiological.shape, neurological.shape)
        width = radiological.shape[0] - 1
        for point in self.points:
            left = ras_to_voxel(radiological.affine, point)
            right = ras_to_voxel(neurological.affine, point)
            with self.subTest(point=point):
                self.assertAlmostEqual(float(left[0]) + float(right[0]), width, places=4)
                # The other two axes are untouched by the preset.
                np.testing.assert_allclose(left[1:], right[1:], atol=1e-6)

    def test_direction_labels_are_derived_from_the_axis_codes(self) -> None:
        # (column axis, row axis) for axial, coronal and sagittal.
        planes = {"axial": (0, 1), "coronal": (0, 2), "sagittal": (1, 2)}
        expected = {
            "radiological": {
                "axial": ("R", "L", "A", "P"),
                "coronal": ("R", "L", "S", "I"),
                "sagittal": ("A", "P", "S", "I"),
            },
            "neurological": {
                "axial": ("L", "R", "A", "P"),
                "coronal": ("L", "R", "S", "I"),
                "sagittal": ("A", "P", "S", "I"),
            },
        }
        for preset, axcodes in ((name, preset_axcodes(name)) for name in ORIENTATION_PRESETS):
            for plane, (column_axis, row_axis) in planes.items():
                with self.subTest(preset=preset, plane=plane):
                    labels = plane_direction_labels(axcodes, column_axis, row_axis)
                    self.assertEqual(labels, expected[preset][plane])
                    left, right, top, bottom = labels
                    self.assertEqual(opposite_axcode(left), right)
                    self.assertEqual(opposite_axcode(top), bottom)

    def test_workbook_findings_land_on_the_same_physical_point(self) -> None:
        case_id = self.path.parent.name
        source = os.environ.get("TEST_SOURCE_XLSX")
        temp_dir = Path(tempfile.mkdtemp(prefix="microbleed_orientation_"))
        try:
            db_path = temp_dir / "review.sqlite"
            initialize_store(Path(source), self.path.parent.parent, db_path)
            targets = list_targets(db_path, case_id, "Orientation QA", 1)
            if not targets:
                self.skipTest(f"No findings recorded for {case_id}.")
            for target in targets[:3]:
                ras = tuple(float(value) for value in target["ras"])
                values = {}
                for preset, volume in self.volumes.items():
                    voxel = ras_to_voxel(volume.affine, ras)
                    index = tuple(int(x) for x in clamp_voxel(voxel, volume.shape))
                    values[preset] = float(volume.data[index])
                with self.subTest(target=target["label"]):
                    self.assertEqual(len(set(values.values())), 1, values)
        finally:
            import shutil

            shutil.rmtree(temp_dir, ignore_errors=True)


class LesionSurfaceTests(unittest.TestCase):
    """The 3D view is only worth looking at if the surface is exact.

    Every case below has a face count that can be worked out on paper, which
    is the point: a renderer that quietly drops or duplicates faces would
    still look like a lesion.
    """

    def test_a_single_voxel_has_its_six_faces(self) -> None:
        from imaging import lesion_surface

        mask = np.zeros((3, 3, 3), dtype=bool)
        mask[1, 1, 1] = True
        quads, normals = lesion_surface(mask, (1.0, 1.0, 1.0))
        self.assertEqual(quads.shape, (6, 4, 3))
        self.assertEqual(
            sorted(map(tuple, normals.tolist())),
            [
                (-1.0, 0.0, 0.0),
                (0.0, -1.0, 0.0),
                (0.0, 0.0, -1.0),
                (0.0, 0.0, 1.0),
                (0.0, 1.0, 0.0),
                (1.0, 0.0, 0.0),
            ],
        )
        # Centred, so the caller can rotate about the origin without knowing
        # where in a 256^3 volume the lesion happened to sit.
        corners = quads.reshape(-1, 3)
        self.assertTrue(np.allclose(corners.mean(axis=0), 0.0))
        self.assertTrue(np.allclose(corners.min(axis=0), -0.5))
        self.assertTrue(np.allclose(corners.max(axis=0), 0.5))

    def test_touching_voxels_do_not_show_the_face_between_them(self) -> None:
        from imaging import lesion_surface

        mask = np.zeros((4, 3, 3), dtype=bool)
        mask[1:3, 1, 1] = True
        quads, _normals = lesion_surface(mask, (1.0, 1.0, 1.0))
        self.assertEqual(len(quads), 10, "6 + 6 minus the shared pair")

    def test_a_hollow_mask_shows_its_inside(self) -> None:
        """A mask with a hole in it is exactly what this view is for."""

        from imaging import lesion_surface

        mask = np.ones((5, 5, 5), dtype=bool)
        mask[2, 2, 2] = False
        quads, _normals = lesion_surface(mask, (1.0, 1.0, 1.0))
        self.assertEqual(len(quads), 6 * 25 + 6, "outer 150 plus the cavity's 6")

    def test_anisotropic_voxels_are_not_drawn_as_cubes(self) -> None:
        """Otherwise a lesion three slices thick looks like a cube.

        Real data here: the SWI a mask is painted on measures 0.5 x 0.5 x 1.0
        mm, so a shape read off an unscaled surface would be wrong by a factor
        of two in one direction.
        """

        from imaging import lesion_surface

        mask = np.zeros((3, 3, 3), dtype=bool)
        mask[1, 1, 1] = True
        quads, _normals = lesion_surface(mask, (0.5, 0.5, 2.0))
        corners = quads.reshape(-1, 3)
        extent = corners.max(axis=0) - corners.min(axis=0)
        self.assertTrue(np.allclose(extent, (0.5, 0.5, 2.0)))

    def test_a_mask_touching_the_volume_edge_is_still_closed(self) -> None:
        from imaging import lesion_surface

        mask = np.zeros((2, 2, 2), dtype=bool)
        mask[0, 0, 0] = True
        quads, _normals = lesion_surface(mask, (1.0, 1.0, 1.0))
        self.assertEqual(len(quads), 6, "the faces against the array edge went missing")

    def test_nothing_to_draw_is_not_an_error(self) -> None:
        from imaging import lesion_surface

        quads, normals = lesion_surface(np.zeros((4, 4, 4), dtype=bool), (1.0, 1.0, 1.0))
        self.assertEqual(quads.shape, (0, 4, 3))
        self.assertEqual(normals.shape, (0, 3))

    def test_a_microbleed_sized_lesion_stays_cheap_to_draw(self) -> None:
        """The budget the software renderer was chosen against.

        A 5 mm lesion on 0.5 mm voxels is about 500 voxels; if that came out
        as thousands of faces, QPainter would be the wrong answer and this
        would have to be OpenGL.
        """

        from imaging import lesion_surface

        grid = np.arange(13) - 6.0
        zz, yy, xx = np.meshgrid(grid, grid, grid, indexing="ij")
        ball = (xx ** 2 + yy ** 2 + zz ** 2) <= 25.0
        quads, normals = lesion_surface(ball, (0.5, 0.5, 0.5))
        self.assertEqual(len(quads), len(normals))
        self.assertGreater(int(ball.sum()), 400)
        self.assertLess(len(quads), 1200, f"{len(quads)} faces for a 5 mm lesion")


class BrainContextTests(unittest.TestCase):
    """The volume the lesion is drawn inside, and the smoothing over it."""

    def test_the_context_cube_undoes_anisotropy(self) -> None:
        """Rotating an anisotropic array by index shears the anatomy.

        The SWI a mask is painted on here is 0.859 x 0.859 x 1.0 mm.  Turned
        forty-five degrees without resampling, a head comes out taller than it
        is, which is a lie about where a lesion sits.
        """

        from imaging import isotropic_context

        # A ball, stored in voxels twice as long on one axis as the others.
        grid = [np.arange(n) - (n - 1) / 2.0 for n in (40, 40, 20)]
        zz, yy, xx = np.meshgrid(*grid, indexing="ij")
        ball = ((zz * 1.0) ** 2 + (yy * 1.0) ** 2 + (xx * 2.0) ** 2) <= 15.0 ** 2
        cube, mm = isotropic_context(ball.astype(np.float32), (1.0, 1.0, 2.0), size=48)

        self.assertEqual(cube.shape, (48, 48, 48))
        self.assertAlmostEqual(mm, 40.0 / 48.0, places=6)
        # A ball resampled to equal-sided voxels measures the same across every
        # axis; before the resample it was half as many voxels on one of them.
        filled = np.argwhere(cube > 0.5)
        extent = filled.max(axis=0) - filled.min(axis=0)
        self.assertLessEqual(int(extent.max() - extent.min()), 2, f"still anisotropic: {extent}")

    def test_a_projection_with_no_rotation_is_a_plain_mean(self) -> None:
        """The one case the answer is known independently."""

        from imaging import project_context

        rng = np.random.default_rng(4)
        cube = rng.random((24, 24, 24), dtype=np.float32)
        flat = project_context(cube, np.eye(3, dtype=np.float32))
        # Rows are volume axis 1, columns axis 0 -- the transpose of a mean
        # taken over the depth axis.
        self.assertTrue(np.allclose(flat, cube.mean(axis=2).T, atol=1e-5))

    def test_the_two_depth_halves_add_up_to_the_whole(self) -> None:
        """The near/far split is what puts the lesion inside the head rather
        than in front of it, so the two slabs have to partition the volume
        exactly -- no voxel counted twice, none dropped."""

        from imaging import project_context

        rng = np.random.default_rng(7)
        cube = rng.random((24, 24, 24), dtype=np.float32)
        rotation = np.eye(3, dtype=np.float32)
        whole = project_context(cube, rotation)
        back = project_context(cube, rotation, near=0.0)
        front = project_context(cube, rotation, far=0.0)
        # 24 samples split 0..11 in front and 11..23 behind by the >= / <=
        # bounds, with the plane at 0 falling in neither half of the line.
        line = np.arange(24, dtype=np.float32) - 11.5
        n_back = int((line >= 0).sum())
        n_front = int((line <= 0).sum())
        self.assertEqual(n_back + n_front, 24, "the halves overlap or leave a gap")
        combined = (back * n_back + front * n_front) / 24.0
        self.assertTrue(np.allclose(combined, whole, atol=1e-5))

    def test_smoothing_keeps_the_surface_it_was_given(self) -> None:
        """It rounds off the stored surface; it does not fetch another one.

        The distinction matters: an isosurface re-derived from an interpolated
        threshold would show a lesion the reader never painted, and this
        window exists to show what is stored.
        """

        from imaging import lesion_surface

        grid = np.arange(13) - 6.0
        zz, yy, xx = np.meshgrid(grid, grid, grid, indexing="ij")
        ball = (xx ** 2 + yy ** 2 + zz ** 2) <= 25.0
        blocky, blocky_normals = lesion_surface(ball, (0.5, 0.5, 0.5))
        smooth, smooth_normals = lesion_surface(ball, (0.5, 0.5, 0.5), smooth=2)

        self.assertEqual(len(smooth), len(blocky), "smoothing added or dropped a face")
        self.assertFalse(np.allclose(smooth, blocky), "smoothing did nothing")
        self.assertTrue(
            np.allclose(np.linalg.norm(smooth_normals, axis=1), 1.0),
            "smoothed normals are not unit vectors",
        )
        # Still pointing outwards, not flipped by the winding.
        self.assertTrue(
            np.all(np.einsum("ij,ij->i", smooth_normals, blocky_normals) > 0),
            "a smoothed face turned inside out",
        )
        # It shrinks, and by how much is documented -- 1.4% on this ball after
        # two passes.  A silent 20% would make the picture a different lesion.
        before = np.linalg.norm(blocky.reshape(-1, 3), axis=1).mean()
        after = np.linalg.norm(smooth.reshape(-1, 3), axis=1).mean()
        self.assertLess(before - after, before * 0.03, "smoothing pulled the surface in too far")
        self.assertGreater(before - after, 0.0)

    def test_the_lesion_can_be_measured_from_the_volume_centre(self) -> None:
        """Drawn inside the head it has to sit where its coordinate says."""

        from imaging import lesion_surface

        mask = np.zeros((20, 20, 20), dtype=bool)
        mask[15, 3, 10] = True
        spacing = (1.0, 1.0, 1.0)
        centre = tuple(value * 20 / 2.0 for value in spacing)
        own, _normals = lesion_surface(mask, spacing)
        placed, _normals = lesion_surface(mask, spacing, centre=centre)
        self.assertTrue(np.allclose(own.reshape(-1, 3).mean(axis=0), 0.0))
        # 15.5 - 10, 3.5 - 10, 10.5 - 10 in voxel centres.
        self.assertTrue(
            np.allclose(placed.reshape(-1, 3).mean(axis=0), (5.5, -6.5, 0.5)),
            placed.reshape(-1, 3).mean(axis=0),
        )


class LabelWriteSafetyTests(unittest.TestCase):
    """One file holds every mask this reader drew for the case."""

    def _reference(self):
        from imaging import Volume

        data = np.zeros((6, 5, 4), dtype=np.float32)
        affine = np.eye(4)
        return Volume(
            path="reference.nii.gz",
            data=data,
            affine=affine,
            shape=data.shape,
            voxel_sizes=(1.0, 1.0, 1.0),
            source_affine=affine,
            source_shape=data.shape,
            source_codes=(1, 1),
        )

    def test_a_failed_write_leaves_the_previous_masks_alone(self) -> None:
        """Written in place, a crash or a full disk halfway through takes
        every finding in the case with it.  Twenty-five of them on the busiest
        case here, and the reader has no way to know until they reopen it."""

        import nibabel as nib

        import imaging

        reference = self._reference()
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "F000001_round1.nii.gz"
            first = np.zeros(reference.shape, dtype=np.uint16)
            first[1:3, 1:3, 1:3] = 7
            imaging.save_label_volume(path, first, reference)
            good = path.read_bytes()

            original = nib.save

            def fails_after_writing(*args, **kwargs):
                original(*args, **kwargs)
                raise OSError("no space left on device")

            nib.save = fails_after_writing
            try:
                with self.assertRaises(OSError):
                    imaging.save_label_volume(path, first * 2, reference)
            finally:
                nib.save = original

            self.assertEqual(path.read_bytes(), good, "the earlier masks were damaged")
            self.assertEqual(
                sorted(p.name for p in Path(folder).iterdir()),
                [path.name],
                "a half-written file was left behind",
            )

    def test_a_completed_write_reads_back(self) -> None:
        import imaging

        reference = self._reference()
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "F000001_round1.nii.gz"
            labels = np.zeros(reference.shape, dtype=np.uint16)
            labels[2:4, 1:4, 0:2] = 3
            imaging.save_label_volume(path, labels, reference)
            stored = imaging.load_volume(path)
            self.assertEqual(tuple(stored.shape), tuple(reference.shape))
            self.assertEqual(int(np.asarray(stored.data).max()), 3)
            self.assertEqual(int((np.asarray(stored.data) > 0).sum()), int((labels > 0).sum()))


class DatasetConfigTests(unittest.TestCase):
    """One study's shape was hard-coded; now it is a file anyone can write.

    These check the file can say what a different study needs, and that it
    refuses what cannot work rather than producing a viewer where every case
    reads as missing.
    """

    def setUp(self) -> None:
        import dataset_config
        import review_store

        self.dataset_config = dataset_config
        self.review_store = review_store
        self.addCleanup(review_store.configure, {})

    def test_an_absent_file_means_the_built_in_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            config = self.dataset_config.load(Path(folder) / "config.json")
        self.assertEqual(config["workbook"]["sheet"], "MCH-microhemorrage")
        self.assertTrue(config["sequences"]["swi"]["required"])

    def test_a_partial_file_keeps_the_rest_of_the_defaults(self) -> None:
        """Nobody should have to restate thirteen column names to rename one."""

        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "config.json"
            path.write_text(
                json.dumps({"workbook": {"columns": {"case_id": "subject"}}}),
                encoding="utf-8",
            )
            config = self.dataset_config.load(path)
        self.assertEqual(config["workbook"]["columns"]["case_id"], "subject")
        self.assertEqual(config["workbook"]["columns"]["ras_l"], "RAS L-R L")
        self.assertEqual(config["workbook"]["sheet"], "MCH-microhemorrage")

    def test_a_configuration_that_cannot_work_is_refused(self) -> None:
        """Each of these produces a viewer that starts and then shows nothing.

        Saying so at load time beats a case list where every row reads
        "missing" and no message says why.
        """

        for broken, why in (
            ({"workbook": {"sheet": ""}}, "no sheet"),
            ({"workbook": {"columns": {"case_id": ""}}}, "no case column"),
            ({"sequences": {"swi": {"suffix": ""}}}, "no filename ending"),
            (
                {"sequences": {k: {"required": False} for k in ("swi", "qsm", "mip")}},
                "nothing required",
            ),
            (
                {
                    "sequences": {
                        "swi": {"required": True, "segmentable": False},
                        "qsm": {"required": False},
                        "mip": {"required": False},
                    }
                },
                "nothing segmentable is required",
            ),
            ({"sequences": {"qsm": {"label": "SWI"}}}, "two sequences share a label"),
            ({"sequences": {"t1": {"suffix": "_T1.nii.gz"}}}, "unknown slot"),
        ):
            with self.subTest(reason=why):
                with self.assertRaises(self.dataset_config.ConfigError):
                    self.dataset_config.validate(broken)

    def test_a_saved_file_reads_back_the_same(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "config.json"
            written = self.dataset_config.validate(
                {
                    "workbook": {"sheet": "Findings", "columns": {"case_id": "subject"}},
                    "sequences": {"mip": {"required": True, "label": "T2*"}},
                }
            )
            self.dataset_config.save(written, path)
            self.assertEqual(self.dataset_config.load(path), written)
            self.assertEqual(
                sorted(p.name for p in Path(folder).iterdir()),
                ["config.json"],
                "a half-written file was left behind",
            )

    def test_the_store_reads_a_workbook_whose_columns_are_named_differently(self) -> None:
        """The point of the exercise: somebody else's spreadsheet."""

        from openpyxl import Workbook

        book = Workbook()
        sheet = book.active
        sheet.title = "Findings"
        sheet.append(["subject", "x", "y", "z", "is_bleed", "where"])
        sheet.append(["CASE01", 1.0, 2.0, 3.0, 1, "pons"])
        sheet.append(["CASE01", -4.0, 5.0, -6.0, 0, "thalamus"])
        sheet.append(["CASE02", 7.5, -8.5, 9.5, 1, ""])

        with tempfile.TemporaryDirectory() as folder:
            workbook_path = Path(folder) / "other_study.xlsx"
            book.save(workbook_path)
            self.review_store.configure(
                {
                    "workbook": {
                        "sheet": "Findings",
                        "columns": {
                            "case_id": "subject",
                            "ras_l": "x",
                            "ras_p": "y",
                            "ras_s": "z",
                            "verify": "is_bleed",
                            "region": "where",
                        },
                    }
                }
            )
            db_path = Path(folder) / "review.sqlite"
            report = initialize_store(workbook_path, Path(folder) / "Data", db_path)
            self.assertEqual(report["source_count"], 3)
            self.assertEqual(report["case_count"], 2)
            targets = list_targets(db_path, "CASE01", "Reader", 1)
            self.assertEqual(len(targets), 2)
            self.assertEqual(tuple(targets[0]["ras"]), (1.0, 2.0, 3.0))
            self.assertEqual(targets[0]["atlasregion"], "pons")
            self.assertEqual(int(targets[0]["source_verify"]), 1)

    def test_a_missing_column_says_what_the_sheet_does_have(self) -> None:
        """"Missing required source columns: subjectid" is unhelpful when the
        column is right there under another name."""

        from openpyxl import Workbook

        book = Workbook()
        sheet = book.active
        sheet.title = self.review_store.MAIN_SHEET
        sheet.append(["subject", "x", "y", "z", "is_bleed"])
        sheet.append(["CASE01", 1.0, 2.0, 3.0, 1])
        with tempfile.TemporaryDirectory() as folder:
            workbook_path = Path(folder) / "mismatched.xlsx"
            book.save(workbook_path)
            with self.assertRaises(self.review_store.SourceReadError) as caught:
                initialize_store(workbook_path, Path(folder) / "Data", Path(folder) / "r.sqlite")
        message = str(caught.exception)
        self.assertIn("subjectid", message, "it should name what it wanted")
        self.assertIn("subject", message, "it should name what the sheet has")
        self.assertIn("config.json", message, "it should say where the fix is")

    def test_an_optional_sequence_does_not_hold_a_case_back(self) -> None:
        """A study with no projection should not have every case incomplete."""

        from openpyxl import Workbook

        book = Workbook()
        sheet = book.active
        sheet.title = "S"
        sheet.append(["subjectid", "RAS L-R L", "RAS P-A A", "RAS I-S S", "verify (yes=1)"])
        sheet.append(["CASE01", 1.0, 2.0, 3.0, 1])

        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            workbook_path = root / "w.xlsx"
            book.save(workbook_path)
            case = root / "Data" / "CASE01"
            case.mkdir(parents=True)
            # Only the two required sequences are on disk.
            (case / "CASE01_swi.nii.gz").write_bytes(b"")
            (case / "CASE01_qsm.nii.gz").write_bytes(b"")
            shape = {
                "workbook": {"sheet": "S"},
                "sequences": {
                    "swi": {"suffix": "_swi.nii.gz"},
                    "qsm": {"suffix": "_qsm.nii.gz"},
                    "mip": {"suffix": "_mip.nii.gz", "required": False},
                },
            }
            self.review_store.configure(shape)
            report = initialize_store(workbook_path, root / "Data", root / "a.sqlite")
            self.assertEqual(report["inventory"]["complete"], 1, "the optional MIP held it back")

            # Demand the projection and the same case is only partial.
            shape["sequences"]["mip"]["required"] = True
            self.review_store.configure(shape)
            report = initialize_store(workbook_path, root / "Data", root / "b.sqlite")
            self.assertEqual(report["inventory"]["complete"], 0)
            self.assertEqual(report["inventory"]["partial"], 1)

    def test_the_filename_endings_can_be_read_off_the_data(self) -> None:
        """Nobody knows their own suffixes by heart; the folder does."""

        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            for case, date in (("A1", "20240101"), ("A2", "20240202"), ("A3", "20240303")):
                case_dir = root / case
                case_dir.mkdir()
                for tail in ("_GRE_SWI_Restored.nii.gz", "_GRE_QSM_Restored.nii.gz"):
                    (case_dir / f"{case}-{date}{tail}").write_bytes(b"")
                (case_dir / f"{case}-{date}_scratch.txt").write_text("", encoding="utf-8")
            found = self.dataset_config.suggest_suffixes(root)
        self.assertIn("_GRE_SWI_Restored.nii.gz", found)
        self.assertIn("_GRE_QSM_Restored.nii.gz", found)
        self.assertTrue(all(item.endswith((".nii", ".nii.gz")) for item in found), found)
        # Longest first: a shorter ending could match two sequences at once.
        self.assertEqual(found, sorted(found, key=len, reverse=True))


if __name__ == "__main__":
    unittest.main()
